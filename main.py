from __future__ import annotations

import argparse
import csv
import getpass
import html
import json
import queue
import re
import shlex
import subprocess
import sys
import threading
import time
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from string import Formatter
from typing import Any
from xml.etree import ElementTree as ET

import clock

# Cooperative run cancellation. The web app assigns a callable here that returns
# True when the user clicks "Stop"; the spoke loops check it between devices and
# halt cleanly (the in-flight device finishes, no new ones start).
_cancel_check = None


def cancellation_requested() -> bool:
    try:
        return bool(_cancel_check and _cancel_check())
    except Exception:
        return False


_PARAMIKO_IMPORT_ERROR: str = ""
try:
    import paramiko as _paramiko_lib
    _PARAMIKO_OK = True
except Exception as _e:
    _paramiko_lib = None  # type: ignore[assignment]
    _PARAMIKO_OK = False
    _PARAMIKO_IMPORT_ERROR = str(_e)

_use_paramiko: bool = False
_paramiko_user: str = ""
_paramiko_pass: str = ""


NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
}

DEFAULT_DELAY_SECONDS = 0
DEFAULT_TRAFFICTEST_PORT = "5201"
DEFAULT_HUB_SERVER_INTF = "Mobily"
DEFAULT_SPOKE_CLIENT_INTF = "wan1"
DEFAULT_HUB_SERVER_START_DELAY_SECONDS = 30.0
# FortiGate's diagnose traffictest run uses a built-in 10s test when no -t is given.
DEFAULT_TRAFFICTEST_DURATION_SECONDS = 10
DEFAULT_SSH_TEMPLATE = 'ssh -o BatchMode=yes -o ConnectTimeout=10 -o StrictHostKeyChecking=accept-new {target} "{remote_command}"'
FORTIGATE_HUB_SETUP_COMMANDS = [
    "diagnose traffictest server-intf {hub_server_intf}",
    "diagnose traffictest port {traffictest_port}",
]
FORTIGATE_HUB_SERVER_COMMAND = "diagnose traffictest run -s"
# Pre-flight: confirm the hub server interface permits speed-test before we start
# the traffictest server on it (otherwise every spoke against this hub would fail).
FORTIGATE_HUB_ALLOWACCESS_CHECK = "show system interface {hub_server_intf}"
# Pre-flight routing checks. On the hub we confirm each spoke IP routes out via the
# server interface; on the spoke we confirm the hub WAN IP routes out via the client
# interface. The routing-table detail output names the egress interface when correct.
FORTIGATE_HUB_ROUTING_CHECK = "get router info routing-table details {spoke_ip}"
FORTIGATE_SPOKE_ROUTING_CHECK = "get router info routing-table details {hub_ip}"
FORTIGATE_SPOKE_COMMANDS = [
    "diagnose traffictest client-intf {spoke_client_intf}",
    "diagnose traffictest port {traffictest_port}",
    # {duration_flag} resolves to " -t <seconds>" when a duration is set for the
    # site, or "" otherwise (FortiGate then uses its built-in 10s default).
    "diagnose traffictest run -b {speed_with_margin} -c {hub_ip}{duration_flag}",
]
THROUGHPUT_PATTERN = re.compile(r"(\d+(?:\.\d+)?)\s*([KMGTP]?bits/sec)", re.IGNORECASE)
ROLE_PATTERN = re.compile(r"\b(sender|receiver)\b", re.IGNORECASE)
# iperf3 summary line: [ 7]  0.00-10.02  sec  4.02 MBytes  3.37 Mbits/sec  72  sender
_IPERF3_SUMMARY_RE = re.compile(
    r"\[\s*\d+\]\s+[\d.]+-[\d.]+\s+sec\s+[\d.]+\s+\S+Bytes\s+"
    r"(\d+(?:\.\d+)?)\s*([KMGTP]?bits/sec)"
    r"(?:\s+(\d+))?\s*(sender|receiver)",
    re.IGNORECASE,
)
NUMBER_PATTERN = re.compile(r"(\d+(?:\.\d+)?)")
ANSI_ESCAPE_PATTERN = re.compile(r"\x1b\[[0-?]*[ -/]*[@-~]")
_FORTIGATE_PROMPT_RE = re.compile(r"[#$]\s*$", re.MULTILINE)
_FORTIGATE_BANNER_RE = re.compile(
    r"press\s+['\"]?a['\"]?|type\s+['\"]?a['\"]?|accept.*disclaimer|disclaimer.*accept"
    r"|you\s+agree|acknowledge|post.?logon",
    re.IGNORECASE,
)
# iperf3 / traffictest errors that FortiGate reports in stdout with exit code 0
_IPERF_ERROR_RE = re.compile(
    r"iperf3?\s*:\s*error|unable\s+to\s+connect|network\s+is\s+unreachable"
    r"|connection\s+refused|connection\s+timed?\s*out|no\s+route\s+to\s+host"
    r"|failed\s+to\s+(connect|send|receive)",
    re.IGNORECASE,
)
# `show system interface <name>` lists permitted services as `set allowaccess ping https ssh speed-test`.
_ALLOWACCESS_RE = re.compile(r"set\s+allowaccess\s+(?P<services>.+)", re.IGNORECASE)
_SPEEDTEST_TOKEN_RE = re.compile(r"\bspeed[-_ ]?test\b", re.IGNORECASE)


def speedtest_allowed(show_output: str) -> bool | None:
    """Whether the interface permits speed-test, parsed from `show system interface`.

    Returns True/False when an ``allowaccess`` line is found, or None when the
    output has no such line (interface not found / unparseable) so the caller can
    decide whether to proceed.
    """
    match = _ALLOWACCESS_RE.search(show_output or "")
    if not match:
        return None
    return bool(_SPEEDTEST_TOKEN_RE.search(match.group("services")))


def routing_via_interface(routing_output: str, interface: str) -> bool:
    """Whether `get router info routing-table details <ip>` resolves via `interface`.

    FortiGate names the egress interface in the resolved route line(s) (e.g.
    ``* 10.255.0.1, via Mobily`` or ``directly connected, wan2``). Returns True only
    when that interface name appears; empty / "no route" / error output is False so
    an unroutable destination is treated as a failed check.
    """
    if not routing_output or not interface:
        return False
    return bool(re.search(r"\b" + re.escape(interface.strip()) + r"\b", routing_output))
FIREWALL_NAME_PATTERNS = [
    re.compile(r"^\s*hostname\s*[:=]\s*(?P<name>.+?)\s*$", re.IGNORECASE),
    re.compile(r"^\s*system\s+name\s*[:=]\s*(?P<name>.+?)\s*$", re.IGNORECASE),
    re.compile(r"^\s*sysname\s*[:=]\s*(?P<name>.+?)\s*$", re.IGNORECASE),
    re.compile(r"^\s*device\s+name\s*[:=]\s*(?P<name>.+?)\s*$", re.IGNORECASE),
    re.compile(r"""^\s*set\s+hostname\s+["']?(?P<name>[^"'\s]+)["']?\s*$""", re.IGNORECASE),
]
NAME_ALIASES = {"name", "site", "site_name", "spoke", "spoke_name", "branch"}
DISCOVERED_NAME_KEYS = NAME_ALIASES | {"firewall_name", "hostname", "device_name"}
IP_ALIASES = {"ip", "host", "address", "spoke_ip", "branch_ip", "wan_ip"}
HUB_IP_ALIASES = {"hub_ip", "hub", "hub_host", "hub_address", "hub_wan_ip"}
HUB_MGMT_IP_ALIASES = {"hub_mgmt_ip", "hub_management_ip", "hub_ssh_ip", "hub_admin_ip", "hub_mgmt"}
SPEED_ALIASES = {
    "speed",
    "rate",
    "bandwidth",
    "expected_speed",
    "speed_mbps",
    "bandwidth_mbps",
}
ACCEPTED_SPEED_ALIASES = {
    "accepted_speed",
    "accepted_speed_mbps",
    "acceptable_speed",
    "min_speed",
    "minimum_speed",
    "threshold_speed",
}
# Fraction of the configured speed a site must reach to pass, used when no
# per-device accepted_speed is set. Overridable per device via accepted_speed.
DEFAULT_ACCEPT_RATIO = 0.90
HUB_SERVER_INTF_ALIASES = {
    "hub_server_intf",
    "server_intf",
    "hub_intf",
    "hub_interface",
    "server_interface",
}
SPOKE_CLIENT_INTF_ALIASES = {
    "spoke_client_intf",
    "client_intf",
    "spoke_intf",
    "spoke_interface",
    "client_interface",
    "wan_intf",
    "wan_interface",
}
TRAFFICTEST_PORT_ALIASES = {
    "traffictest_port",
    "traffic_port",
    "iperf_port",
    "test_port",
}
TRAFFICTEST_DURATION_ALIASES = {
    "traffictest_duration",
    "test_duration",
    "duration",
    "duration_seconds",
    "traffic_duration",
}
CIRCUIT_ID_ALIASES = {
    "circuit_id",
    "circuit",
    "circuitid",
    "circuit_no",
    "circuit_number",
    "circuit_reference",
    "circuit_ref",
}
ISP_ALIASES = {
    "isp",
    "carrier",
    "provider",
    "service_provider",
    "telco",
}


@dataclass
class SiteDefinition:
    index: int
    raw: dict[str, str]
    placeholders: dict[str, str]
    display_name: str
    ip_address: str
    hub_ip: str
    speed: str
    speed_mbps: float | None
    speed_with_margin_mbps: float | None
    speed_with_margin_label: str
    accepted_speed: str = ""
    accepted_speed_mbps: float | None = None
    hub_mgmt_ip: str = ""
    hub_server_intf: str = ""
    spoke_client_intf: str = ""
    traffictest_port: str = ""
    traffictest_duration: str = ""
    hub_name: str = ""
    circuit_id: str = ""
    isp: str = ""


@dataclass
class CommandResult:
    template: str
    command: str
    started_at: datetime
    ended_at: datetime
    return_code: int | None
    stdout: str
    stderr: str
    error: str | None = None
    throughput_mbps: float | None = None
    throughput_label: str | None = None
    retransmissions: int | None = None
    sender_throughput_mbps: float | None = None
    receiver_throughput_mbps: float | None = None

    @property
    def duration_seconds(self) -> float:
        return (self.ended_at - self.started_at).total_seconds()

    @property
    def status(self) -> str:
        if self.error:
            return "template-error"
        if self.return_code == 0:
            if _IPERF_ERROR_RE.search(self.stdout) or _IPERF_ERROR_RE.search(self.stderr):
                return "failed"
            return "success"
        return "failed"


@dataclass
class SiteRun:
    site: SiteDefinition
    started_at: datetime
    ended_at: datetime
    command_results: list[CommandResult] = field(default_factory=list)
    name_discovery_result: CommandResult | None = None
    delayed_after_seconds: int = 0

    @property
    def duration_seconds(self) -> float:
        return (self.ended_at - self.started_at).total_seconds()

    @property
    def status(self) -> str:
        if not self.command_results:
            if self.name_discovery_result and self.name_discovery_result.error:
                return "failed"
            return "skipped"
        if any(result.status != "success" for result in self.command_results):
            return "failed"
        return "success"

    @property
    def max_throughput_mbps(self) -> float | None:
        values = [r.throughput_mbps for r in self.command_results if r.throughput_mbps is not None]
        return max(values) if values else None

    @property
    def max_sender_throughput_mbps(self) -> float | None:
        values = [r.sender_throughput_mbps for r in self.command_results if r.sender_throughput_mbps is not None]
        return max(values) if values else None

    @property
    def max_receiver_throughput_mbps(self) -> float | None:
        values = [r.receiver_throughput_mbps for r in self.command_results if r.receiver_throughput_mbps is not None]
        return max(values) if values else None


class SafeFormatDict(dict[str, str]):
    def __missing__(self, key: str) -> str:
        raise KeyError(key)


def sanitize_key(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = value.strip("_")
    return value or "column"


def column_letters_to_index(reference: str) -> int:
    letters = "".join(character for character in reference if character.isalpha())
    index = 0
    for character in letters:
        index = (index * 26) + (ord(character.upper()) - ord("A") + 1)
    return index - 1


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for item in root.findall("main:si", NS):
        parts = [node.text or "" for node in item.findall(".//main:t", NS)]
        strings.append("".join(parts))
    return strings


def resolve_sheet_path(archive: zipfile.ZipFile, sheet_name: str | None) -> tuple[str, str]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        relation.attrib["Id"]: relation.attrib["Target"]
        for relation in rel_root.findall("pkg:Relationship", NS)
    }

    sheets = workbook_root.findall("main:sheets/main:sheet", NS)
    if not sheets:
        raise ValueError("No worksheets found in the Excel file.")

    selected = None
    if sheet_name:
        for sheet in sheets:
            if sheet.attrib.get("name") == sheet_name:
                selected = sheet
                break
        if selected is None:
            available = ", ".join(sheet.attrib.get("name", "<unnamed>") for sheet in sheets)
            raise ValueError(f"Worksheet '{sheet_name}' not found. Available sheets: {available}")
    else:
        selected = sheets[0]

    relationship_id = selected.attrib[f"{{{NS['rel']}}}id"]
    target = rel_map[relationship_id].lstrip("/")
    if not target.startswith("xl/"):
        target = f"xl/{target}"
    return target, selected.attrib.get("name", "Sheet1")


def extract_cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//main:t", NS))

    value_node = cell.find("main:v", NS)
    if value_node is None or value_node.text is None:
        return ""

    raw_value = value_node.text
    if cell_type == "s":
        return shared_strings[int(raw_value)]
    return raw_value


def load_xlsx_rows(path: Path, sheet_name: str | None) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = load_shared_strings(archive)
        sheet_path, _ = resolve_sheet_path(archive, sheet_name)
        sheet_root = ET.fromstring(archive.read(sheet_path))

    rows: list[dict[int, str]] = []
    for row in sheet_root.findall(".//main:sheetData/main:row", NS):
        parsed_row: dict[int, str] = {}
        for cell in row.findall("main:c", NS):
            reference = cell.attrib.get("r", "")
            column_index = column_letters_to_index(reference)
            parsed_row[column_index] = extract_cell_value(cell, shared_strings).strip()
        if any(value for value in parsed_row.values()):
            rows.append(parsed_row)

    if not rows:
        return []

    header_row = rows[0]
    max_index = max(header_row)
    headers = [header_row.get(index, "").strip() for index in range(max_index + 1)]

    records: list[dict[str, str]] = []
    for row in rows[1:]:
        record: dict[str, str] = {}
        for index, header in enumerate(headers):
            if header:
                record[header] = row.get(index, "").strip()
        if any(record.values()):
            records.append(record)
    return records


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [{key.strip(): (value or "").strip() for key, value in row.items()} for row in reader]


def load_rows(path: Path, sheet_name: str | None) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv_rows(path)
    if suffix == ".xlsx":
        return load_xlsx_rows(path, sheet_name)
    raise ValueError("Only .csv and .xlsx input files are supported.")


def find_first_value(placeholders: dict[str, str], aliases: set[str]) -> str:
    for alias in aliases:
        if placeholders.get(alias):
            return placeholders[alias]
    return ""


def parse_speed_to_mbps(raw_speed: str) -> float | None:
    if not raw_speed:
        return None

    cleaned = raw_speed.strip().replace(",", "")
    match = NUMBER_PATTERN.search(cleaned)
    if not match:
        return None

    value = float(match.group(1))
    normalized = cleaned.lower().replace(" ", "")

    if any(unit in normalized for unit in ("gbps", "gbit", "gbits")) or normalized.endswith("g"):
        return value * 1000
    if any(unit in normalized for unit in ("kbps", "kbit", "kbits")) or normalized.endswith("k"):
        return value / 1000
    # Megabits per second is the base unit. Check it before the bare-bits check
    # below, since "bps"/"bit" are substrings of "mbps"/"mbit".
    if any(unit in normalized for unit in ("mbps", "mbit", "mbits")) or normalized.endswith("m"):
        return value
    if any(unit in normalized for unit in ("bps", "bit", "bits")):
        return value / 1_000_000
    return value


def format_mbps_for_traffictest(speed_mbps: float | None) -> str:
    if speed_mbps is None:
        return ""

    rounded = round(speed_mbps, 2)
    if rounded.is_integer():
        return f"{int(rounded)}M"
    return f"{rounded:g}M"


def _sanitize_duration(value) -> str:
    """Return a positive-integer seconds string, or "" when absent/invalid."""
    try:
        seconds = int(str(value).strip())
    except (TypeError, ValueError):
        return ""
    return str(seconds) if seconds > 0 else ""


def build_sites(rows: list[dict[str, str]]) -> list[SiteDefinition]:
    sites: list[SiteDefinition] = []
    for index, row in enumerate(rows, start=1):
        placeholders: dict[str, str] = {}
        for key, value in row.items():
            sanitized = sanitize_key(key)
            if sanitized in placeholders and placeholders[sanitized]:
                continue
            placeholders[sanitized] = value

        ip_address = find_first_value(placeholders, IP_ALIASES)
        hub_ip = find_first_value(placeholders, HUB_IP_ALIASES)
        hub_mgmt_ip = find_first_value(placeholders, HUB_MGMT_IP_ALIASES)
        display_name = ip_address or f"spoke-{index}"
        for key in DISCOVERED_NAME_KEYS:
            placeholders[key] = display_name
        if hub_ip:
            placeholders.setdefault("hub_ip", hub_ip)
            placeholders.setdefault("hub", hub_ip)
        if hub_mgmt_ip:
            placeholders.setdefault("hub_mgmt_ip", hub_mgmt_ip)
        speed = find_first_value(placeholders, SPEED_ALIASES)
        accepted_speed = find_first_value(placeholders, ACCEPTED_SPEED_ALIASES)
        hub_server_intf = find_first_value(placeholders, HUB_SERVER_INTF_ALIASES)
        spoke_client_intf = find_first_value(placeholders, SPOKE_CLIENT_INTF_ALIASES)
        traffictest_port = find_first_value(placeholders, TRAFFICTEST_PORT_ALIASES)
        traffictest_duration = find_first_value(placeholders, TRAFFICTEST_DURATION_ALIASES)
        circuit_id = find_first_value(placeholders, CIRCUIT_ID_ALIASES)
        isp = find_first_value(placeholders, ISP_ALIASES)
        speed_mbps = parse_speed_to_mbps(speed)
        accepted_speed_mbps = parse_speed_to_mbps(accepted_speed)
        speed_with_margin_mbps = round(speed_mbps * 1.15, 2) if speed_mbps is not None else None
        speed_with_margin_label = format_mbps_for_traffictest(speed_with_margin_mbps)

        if speed_mbps is not None:
            placeholders.setdefault("speed_mbps", f"{speed_mbps:g}")
        if speed_with_margin_mbps is not None:
            placeholders.setdefault("speed_with_margin_mbps", f"{speed_with_margin_mbps:g}")
        if speed_with_margin_label:
            placeholders.setdefault("speed_with_margin", speed_with_margin_label)
            placeholders.setdefault("bandwidth_with_margin", speed_with_margin_label)

        sites.append(
            SiteDefinition(
                index=index,
                raw=row,
                placeholders=placeholders,
                display_name=display_name,
                ip_address=ip_address,
                hub_ip=hub_ip,
                hub_mgmt_ip=hub_mgmt_ip,
                speed=speed,
                speed_mbps=speed_mbps,
                accepted_speed=accepted_speed,
                accepted_speed_mbps=accepted_speed_mbps,
                speed_with_margin_mbps=speed_with_margin_mbps,
                speed_with_margin_label=speed_with_margin_label,
                hub_server_intf=hub_server_intf,
                spoke_client_intf=spoke_client_intf,
                traffictest_port=traffictest_port,
                traffictest_duration=traffictest_duration,
                circuit_id=circuit_id,
                isp=isp,
            )
        )
    return sites


def load_command_templates(args: argparse.Namespace) -> list[str]:
    templates = list(args.command or [])
    if args.command_file:
        file_path = Path(args.command_file)
        with file_path.open("r", encoding="utf-8") as handle:
            templates.extend(
                line.strip()
                for line in handle
                if line.strip() and not line.lstrip().startswith("#")
            )
    templates = [template for template in templates if template]
    return templates


def fortigate_traffictest_templates() -> list[str]:
    return FORTIGATE_HUB_SETUP_COMMANDS + [FORTIGATE_HUB_SERVER_COMMAND] + FORTIGATE_SPOKE_COMMANDS


def validate_template_fields(templates: list[str], available_keys: set[str]) -> None:
    missing_fields: set[str] = set()
    formatter = Formatter()
    for template in templates:
        for _, field_name, _, _ in formatter.parse(template):
            if field_name and field_name not in available_keys:
                missing_fields.add(field_name)
    if missing_fields:
        available = ", ".join(sorted(available_keys))
        missing = ", ".join(sorted(missing_fields))
        raise ValueError(f"Missing columns/placeholders: {missing}. Available placeholders: {available}")


def extract_throughput(output: str) -> tuple[float | None, str | None, int | None, float | None, float | None]:
    """Return (throughput_mbps, label, retransmissions, sender_mbps, receiver_mbps)."""
    multipliers = {
        "bits/sec": 1 / 1_000_000,
        "kbits/sec": 1 / 1_000,
        "mbits/sec": 1.0,
        "gbits/sec": 1_000.0,
        "tbits/sec": 1_000_000.0,
        "pbits/sec": 1_000_000_000.0,
    }

    summary_matches = list(_IPERF3_SUMMARY_RE.finditer(output))
    if summary_matches:
        receiver_m = next((m for m in summary_matches if m.group(4).lower() == "receiver"), None)
        sender_m = next((m for m in summary_matches if m.group(4).lower() == "sender"), None)
        sender_mbps = float(sender_m.group(1)) * multipliers[sender_m.group(2).lower()] if sender_m else None
        receiver_mbps = float(receiver_m.group(1)) * multipliers[receiver_m.group(2).lower()] if receiver_m else None
        chosen = receiver_m or sender_m or summary_matches[-1]
        throughput_mbps = receiver_mbps or sender_mbps
        label_parts = []
        if sender_mbps is not None:
            label_parts.append(f"{float(sender_m.group(1)):g} {sender_m.group(2)} (sender)")
        if receiver_mbps is not None:
            label_parts.append(f"{float(receiver_m.group(1)):g} {receiver_m.group(2)} (receiver)")
        label = " / ".join(label_parts) if label_parts else f"{float(chosen.group(1)):g} {chosen.group(2)}"
        retransmissions = int(sender_m.group(3)) if sender_m and sender_m.group(3) else None
        return throughput_mbps, label, retransmissions, sender_mbps, receiver_mbps

    # Fall back to last generic bits/sec match (non-iperf3 output)
    matches = list(THROUGHPUT_PATTERN.finditer(output))
    if not matches:
        return None, None, None, None, None
    chosen = matches[-1]
    value = float(chosen.group(1))
    unit = chosen.group(2).lower()
    throughput_mbps = value * multipliers[unit]
    nearby_text = output[max(0, chosen.start() - 60): min(len(output), chosen.end() + 60)]
    role_match = ROLE_PATTERN.search(nearby_text)
    label = f"{value:g} {chosen.group(2)}"
    if role_match:
        label = f"{label} ({role_match.group(1).lower()})"
    return throughput_mbps, label, None, None, None


def clean_firewall_name(value: str) -> str | None:
    cleaned = ANSI_ESCAPE_PATTERN.sub("", value).strip().strip("'\"")
    cleaned = re.sub(r"\s*[>#]\s*$", "", cleaned).strip()
    if not cleaned:
        return None

    lowered = cleaned.lower()
    ignored_prefixes = (
        "warning:",
        "the authenticity of host",
        "permanently added",
        "pseudo-terminal",
        "last login",
        "welcome",
        "ssh:",
        "connection ",
        "permission denied",
        "command not found",
        "bash:",
        "zsh:",
    )
    if any(lowered.startswith(prefix) for prefix in ignored_prefixes):
        return None
    if len(cleaned) > 128:
        return None
    return cleaned


def parse_firewall_name(output: str) -> str | None:
    lines = [ANSI_ESCAPE_PATTERN.sub("", line).strip() for line in output.splitlines()]
    non_empty_lines = [line for line in lines if line]

    for line in non_empty_lines:
        for pattern in FIREWALL_NAME_PATTERNS:
            match = pattern.match(line)
            if match:
                candidate = clean_firewall_name(match.group("name"))
                if candidate:
                    return candidate

    if len(non_empty_lines) == 1:
        return clean_firewall_name(non_empty_lines[0])

    for line in non_empty_lines:
        candidate = clean_firewall_name(line)
        if candidate and re.fullmatch(r"[A-Za-z0-9_.:-]+", candidate):
            return candidate
    return None


def run_command(command: str, timeout: int | None, dry_run: bool) -> CommandResult:
    started_at = clock.now()
    if dry_run:
        ended_at = clock.now()
        return CommandResult(
            template=command,
            command=command,
            started_at=started_at,
            ended_at=ended_at,
            return_code=0,
            stdout="dry-run: command not executed",
            stderr="",
        )

    completed = subprocess.run(
        command,
        shell=True,
        text=True,
        capture_output=True,
        timeout=timeout,
    )
    ended_at = clock.now()
    combined_output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    throughput_mbps, throughput_label, retransmissions, sender_mbps, receiver_mbps = extract_throughput(combined_output)
    return CommandResult(
        template=command,
        command=command,
        started_at=started_at,
        ended_at=ended_at,
        return_code=completed.returncode,
        stdout=completed.stdout,
        stderr=completed.stderr,
        throughput_mbps=throughput_mbps,
        throughput_label=throughput_label,
        retransmissions=retransmissions,
        sender_throughput_mbps=sender_mbps,
        receiver_throughput_mbps=receiver_mbps,
    )


def build_ssh_command(
    site: SiteDefinition,
    ssh_template: str,
    target: str,
    remote_template: str,
) -> tuple[str, str]:
    remote_command = render_template(remote_template, site)
    command = render_template(
        ssh_template,
        site,
        {"target": target, "remote_command": remote_command},
    )
    return remote_template, command


def build_ssh_command_or_error(
    site: SiteDefinition,
    ssh_template: str,
    target: str,
    remote_template: str,
) -> tuple[str | None, CommandResult | None]:
    try:
        _, command = build_ssh_command(site, ssh_template, target, remote_template)
    except KeyError as error:
        moment = clock.now()
        return None, CommandResult(
            template=remote_template,
            command="",
            started_at=moment,
            ended_at=moment,
            return_code=None,
            stdout="",
            stderr="",
            error=f"Missing placeholder value for '{error.args[0]}'",
        )
    return command, None


def run_rendered_command(
    template: str,
    command: str,
    timeout: int | None,
    dry_run: bool,
) -> CommandResult:
    try:
        result = run_command(command, timeout=timeout, dry_run=dry_run)
    except subprocess.TimeoutExpired as error:
        ended_at = clock.now()
        return CommandResult(
            template=template,
            command=command,
            started_at=ended_at,
            ended_at=ended_at,
            return_code=None,
            stdout=error.stdout or "",
            stderr=error.stderr or "",
            error=f"Timed out after {error.timeout} seconds",
        )

    result.template = template
    return result


def start_background_command(
    template: str,
    command: str,
    dry_run: bool,
) -> tuple[CommandResult, subprocess.Popen[str] | None]:
    started_at = clock.now()
    if dry_run:
        return (
            CommandResult(
                template=template,
                command=command,
                started_at=started_at,
                ended_at=clock.now(),
                return_code=0,
                stdout="dry-run: background command not executed",
                stderr="",
            ),
            None,
        )

    process = subprocess.Popen(
        command,
        shell=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    return (
        CommandResult(
            template=template,
            command=command,
            started_at=started_at,
            ended_at=started_at,
            return_code=None,
            stdout="Hub server command started in the background.",
            stderr="",
        ),
        process,
    )


def finalize_background_command(
    initial_result: CommandResult,
    process: subprocess.Popen[str],
    stop_if_running: bool,
) -> CommandResult:
    stopped_after_test = False
    if process.poll() is None and stop_if_running:
        process.terminate()
        stopped_after_test = True

    try:
        stdout, stderr = process.communicate(timeout=5)
    except subprocess.TimeoutExpired:
        process.kill()
        stdout, stderr = process.communicate()
        stopped_after_test = True

    ended_at = clock.now()
    return_code = 0 if stopped_after_test else process.returncode
    if stopped_after_test:
        stdout = "\n".join(
            part
            for part in [
                stdout.strip(),
                "Hub server process stopped after spoke traffictest commands.",
            ]
            if part
        )

    combined_output = "\n".join(part for part in [stdout, stderr] if part)
    throughput_mbps, throughput_label, retransmissions, sender_mbps, receiver_mbps = extract_throughput(combined_output)
    return CommandResult(
        template=initial_result.template,
        command=initial_result.command,
        started_at=initial_result.started_at,
        ended_at=ended_at,
        return_code=return_code,
        stdout=stdout,
        stderr=stderr,
        throughput_mbps=throughput_mbps,
        throughput_label=throughput_label,
        retransmissions=retransmissions,
        sender_throughput_mbps=sender_mbps,
        receiver_throughput_mbps=receiver_mbps,
    )


class _ParamikoHandle:
    """Drop-in for subprocess.Popen so finalize_background_command works unchanged."""

    def __init__(self) -> None:
        self._stdout_parts: list[str] = []
        self._stderr_parts: list[str] = []
        self._done = threading.Event()
        self._shell: Any = None
        self._client: Any = None
        self.returncode: int | None = None

    def poll(self) -> int | None:
        return self.returncode if self._done.is_set() else None

    def terminate(self) -> None:
        for obj in (self._shell, self._client):
            if obj is not None:
                try:
                    obj.close()
                except Exception:
                    pass

    def kill(self) -> None:
        self.terminate()

    def communicate(self, timeout: float | None = None) -> tuple[str, str]:
        self._done.wait(timeout=timeout)
        return "".join(self._stdout_parts), "".join(self._stderr_parts)


def _paramiko_connect(host: str, timeout: int | None) -> Any:
    client = _paramiko_lib.SSHClient()
    client.set_missing_host_key_policy(_paramiko_lib.AutoAddPolicy())
    client.connect(host, username=_paramiko_user, password=_paramiko_pass, timeout=timeout or 10)
    transport = client.get_transport()
    if transport is not None:
        transport.set_keepalive(30)  # send SSH keepalive every 30 s to prevent idle disconnect
    return client


def _shell_recv_chunk(shell: Any) -> str:
    if shell.recv_ready():
        return shell.recv(4096).decode(errors="replace")
    return ""


def _shell_read_until_prompt(shell: Any, timeout: float | None) -> str:
    output = ""
    deadline = (time.time() + timeout) if timeout is not None else None
    while deadline is None or time.time() < deadline:
        chunk = _shell_recv_chunk(shell)
        if chunk:
            output += chunk
            if _FORTIGATE_PROMPT_RE.search(ANSI_ESCAPE_PATTERN.sub("", output)):
                return output
        else:
            time.sleep(0.05)
    return output


def _paramiko_open_shell(client: Any, timeout: int | None) -> Any:
    """Open an interactive FortiGate shell and accept the post-logon banner if present."""
    shell = client.invoke_shell(width=220, height=50)
    shell.settimeout(timeout or 30)
    # Collect initial output for up to 5 s (banner, MOTD, etc.)
    initial = ""
    deadline = time.time() + 5
    while time.time() < deadline:
        chunk = _shell_recv_chunk(shell)
        if chunk:
            initial += chunk
        else:
            time.sleep(0.1)
        clean = ANSI_ESCAPE_PATTERN.sub("", initial)
        if _FORTIGATE_PROMPT_RE.search(clean):
            return shell  # prompt already visible — no banner
        if _FORTIGATE_BANNER_RE.search(clean):
            # Banner requires pressing 'a' to accept
            shell.send("a\n")
            _shell_read_until_prompt(shell, timeout=timeout or 30)
            return shell
    return shell


def _paramiko_exec(
    host: str,
    remote_command: str,
    template: str,
    timeout: int | None,
    dry_run: bool,
) -> CommandResult:
    cmd_label = f"[paramiko] {_paramiko_user}@{host}: {remote_command}"
    started_at = clock.now()
    if dry_run:
        return CommandResult(
            template=template, command=cmd_label,
            started_at=started_at, ended_at=clock.now(),
            return_code=0, stdout="dry-run: command not executed", stderr="",
        )
    try:
        client = _paramiko_connect(host, timeout)
        shell = _paramiko_open_shell(client, timeout)
        shell.send(remote_command + "\n")
        raw = _shell_read_until_prompt(shell, timeout=timeout or 60)
        shell.close()
        client.close()
        stdout = ANSI_ESCAPE_PATTERN.sub("", raw)
        rc = 0
    except Exception as exc:
        ended_at = clock.now()
        return CommandResult(
            template=template, command=cmd_label,
            started_at=started_at, ended_at=ended_at,
            return_code=None, stdout="", stderr="", error=str(exc),
        )
    ended_at = clock.now()
    throughput_mbps, throughput_label, retransmissions, sender_mbps, receiver_mbps = extract_throughput(stdout)
    return CommandResult(
        template=template, command=cmd_label,
        started_at=started_at, ended_at=ended_at,
        return_code=rc, stdout=stdout, stderr="",
        throughput_mbps=throughput_mbps, throughput_label=throughput_label,
        retransmissions=retransmissions,
        sender_throughput_mbps=sender_mbps,
        receiver_throughput_mbps=receiver_mbps,
    )


def _paramiko_start_background(
    host: str,
    remote_command: str,
    template: str,
    dry_run: bool,
) -> tuple[CommandResult, _ParamikoHandle | None]:
    cmd_label = f"[paramiko] {_paramiko_user}@{host}: {remote_command}"
    started_at = clock.now()
    if dry_run:
        return (
            CommandResult(
                template=template, command=cmd_label,
                started_at=started_at, ended_at=clock.now(),
                return_code=0, stdout="dry-run: background command not executed", stderr="",
            ),
            None,
        )
    handle = _ParamikoHandle()

    def _run() -> None:
        try:
            client = _paramiko_connect(host, timeout=None)
            shell = _paramiko_open_shell(client, timeout=30)
            handle._shell = shell
            handle._client = client
            shell.send(remote_command + "\n")
            while not getattr(shell, "closed", False):
                chunk = _shell_recv_chunk(shell)
                if chunk:
                    handle._stdout_parts.append(chunk)
                else:
                    time.sleep(0.1)
        except Exception as exc:
            handle._stderr_parts.append(str(exc))
            handle.returncode = 1
        finally:
            handle.returncode = handle.returncode if handle.returncode is not None else 0
            handle._done.set()

    threading.Thread(target=_run, daemon=True).start()
    return (
        CommandResult(
            template=template, command=cmd_label,
            started_at=started_at, ended_at=started_at,
            return_code=None, stdout="Hub server command started via Paramiko.", stderr="",
        ),
        handle,
    )


def _paramiko_hub_routing_check(
    ssh_target: str,
    spoke_sites: list[SiteDefinition],
    server_intf: str,
    timeout: int | None,
    dry_run: bool,
) -> tuple[list[CommandResult], dict[str, bool]]:
    """On the hub, confirm each spoke IP routes out via the server interface.

    Runs ``get router info routing-table details <spoke_ip>`` at the root prompt
    (before ``config global``) for every spoke that will test against this hub.
    Returns (per-spoke CommandResults, {spoke_ip: routing_ok}).
    """
    cmd_prefix = f"[paramiko] {_paramiko_user}@{ssh_target}: "
    results: list[CommandResult] = []
    verdicts: dict[str, bool] = {}

    if dry_run:
        now = clock.now()
        for site in spoke_sites:
            spoke_ip = site.ip_address or ""
            results.append(CommandResult(
                template=FORTIGATE_HUB_ROUTING_CHECK,
                command=cmd_prefix + f"get router info routing-table details {spoke_ip}",
                started_at=now, ended_at=now, return_code=0,
                stdout="dry-run: routing check not executed", stderr="",
            ))
            verdicts[spoke_ip] = True
        return results, verdicts

    try:
        client = _paramiko_connect(ssh_target, timeout)
        shell = _paramiko_open_shell(client, timeout)
    except Exception as exc:
        now = clock.now()
        results.append(CommandResult(
            template=FORTIGATE_HUB_ROUTING_CHECK, command=cmd_prefix + "(connect)",
            started_at=now, ended_at=now, return_code=None, stdout="", stderr="", error=str(exc),
        ))
        for site in spoke_sites:
            verdicts[site.ip_address or ""] = False
        return results, verdicts

    try:
        for site in spoke_sites:
            spoke_ip = site.ip_address or ""
            cmd = f"get router info routing-table details {spoke_ip}"
            started_at = clock.now()
            try:
                shell.send(cmd + "\n")
                raw = _shell_read_until_prompt(shell, timeout=timeout or 30)
                stdout = ANSI_ESCAPE_PATTERN.sub("", raw)
                ok = routing_via_interface(stdout, server_intf)
                verdicts[spoke_ip] = ok
                results.append(CommandResult(
                    template=FORTIGATE_HUB_ROUTING_CHECK, command=cmd_prefix + cmd,
                    started_at=started_at, ended_at=clock.now(),
                    return_code=0 if ok else None, stdout=stdout, stderr="",
                    error=None if ok else (
                        f"Spoke {spoke_ip} does not route via the hub server interface "
                        f"'{server_intf}' — skipping this spoke."),
                ))
            except Exception as exc:
                verdicts[spoke_ip] = False
                results.append(CommandResult(
                    template=FORTIGATE_HUB_ROUTING_CHECK, command=cmd_prefix + cmd,
                    started_at=started_at, ended_at=clock.now(),
                    return_code=None, stdout="", stderr="", error=str(exc),
                ))
    finally:
        try:
            shell.close(); client.close()
        except Exception:
            pass
    return results, verdicts


def _paramiko_hub_session(
    ssh_target: str,
    rep_site: SiteDefinition,
    setup_templates: list[str],
    server_template: str,
    timeout: int | None,
    dry_run: bool,
    check_speedtest: bool = True,
) -> tuple[list[CommandResult], CommandResult, "_ParamikoHandle | None"]:
    """Run all hub commands in one Paramiko shell, entering global VDOM once."""
    cmd_prefix = f"[paramiko] {_paramiko_user}@{ssh_target}: "

    if dry_run:
        now = clock.now()
        setup_results = []
        if check_speedtest:
            setup_results.append(CommandResult(
                template=FORTIGATE_HUB_ALLOWACCESS_CHECK,
                command=cmd_prefix + render_template(FORTIGATE_HUB_ALLOWACCESS_CHECK, rep_site),
                started_at=now, ended_at=now,
                return_code=0, stdout="dry-run: speed-test allowaccess check not executed", stderr="",
            ))
        setup_results += [
            CommandResult(
                template=t, command=cmd_prefix + render_template(t, rep_site),
                started_at=now, ended_at=now,
                return_code=0, stdout="dry-run: command not executed", stderr="",
            )
            for t in setup_templates
        ]
        server_cmd_str = render_template(server_template, rep_site)
        server_initial = CommandResult(
            template=server_template, command=cmd_prefix + server_cmd_str,
            started_at=now, ended_at=now,
            return_code=0, stdout="dry-run: background command not executed", stderr="",
        )
        return setup_results, server_initial, None

    try:
        client = _paramiko_connect(ssh_target, timeout)
        shell = _paramiko_open_shell(client, timeout)
    except Exception as exc:
        now = clock.now()
        err_result = CommandResult(
            template=setup_templates[0] if setup_templates else server_template,
            command=cmd_prefix + "(connect)",
            started_at=now, ended_at=now,
            return_code=None, stdout="", stderr="", error=str(exc),
        )
        return [err_result], err_result, None

    # Enter global VDOM once for the entire session.
    shell.send("config global\n")
    _shell_read_until_prompt(shell, timeout=timeout or 30)

    setup_results: list[CommandResult] = []
    connection_failed = False

    # Pre-flight: confirm the hub server interface permits speed-test in its
    # allowaccess. If it does not, the traffictest server is unreachable and every
    # spoke against this hub would fail — so stop here with a clear, actionable error.
    if check_speedtest:
        intf = render_template("{hub_server_intf}", rep_site)
        check_rendered = render_template(FORTIGATE_HUB_ALLOWACCESS_CHECK, rep_site)
        started_at = clock.now()
        try:
            shell.send(check_rendered + "\n")
            raw = _shell_read_until_prompt(shell, timeout=timeout or 30)
            stdout = ANSI_ESCAPE_PATTERN.sub("", raw)
            allowed = speedtest_allowed(stdout)
            ended_at = clock.now()
            if allowed is False:
                setup_results.append(CommandResult(
                    template=FORTIGATE_HUB_ALLOWACCESS_CHECK, command=cmd_prefix + check_rendered,
                    started_at=started_at, ended_at=ended_at,
                    return_code=None, stdout=stdout, stderr="",
                    error=(f"Hub interface '{intf}' does not permit speed-test in allowaccess. "
                           f"Enable it on the hub:  config system interface / edit {intf} / "
                           f"append allowaccess speed-test / end"),
                ))
                try:
                    shell.close(); client.close()
                except Exception:
                    pass
                server_initial = CommandResult(
                    template=server_template, command=cmd_prefix + server_template,
                    started_at=clock.now(), ended_at=clock.now(),
                    return_code=None, stdout="", stderr="",
                    error="Skipped — hub setup failed.",
                )
                return setup_results, server_initial, None
            note = (f"speed-test is permitted on '{intf}'." if allowed
                    else f"Could not read allowaccess for '{intf}'; proceeding without the speed-test gate.")
            setup_results.append(CommandResult(
                template=FORTIGATE_HUB_ALLOWACCESS_CHECK, command=cmd_prefix + check_rendered,
                started_at=started_at, ended_at=ended_at,
                return_code=0, stdout=note + "\n\n" + stdout, stderr="",
            ))
        except Exception as exc:
            ended_at = clock.now()
            setup_results.append(CommandResult(
                template=FORTIGATE_HUB_ALLOWACCESS_CHECK, command=cmd_prefix + check_rendered,
                started_at=started_at, ended_at=ended_at,
                return_code=None, stdout="", stderr="", error=str(exc),
            ))
            try:
                shell.close(); client.close()
            except Exception:
                pass
            server_initial = CommandResult(
                template=server_template, command=cmd_prefix + server_template,
                started_at=clock.now(), ended_at=clock.now(),
                return_code=None, stdout="", stderr="", error="Skipped — hub setup failed.",
            )
            return setup_results, server_initial, None

    for template in setup_templates:
        rendered = render_template(template, rep_site)
        cmd_label = cmd_prefix + rendered
        started_at = clock.now()
        try:
            shell.send(rendered + "\n")
            raw = _shell_read_until_prompt(shell, timeout=timeout or 60)
            stdout = ANSI_ESCAPE_PATTERN.sub("", raw)
            ended_at = clock.now()
            setup_results.append(CommandResult(
                template=template, command=cmd_label,
                started_at=started_at, ended_at=ended_at,
                return_code=0, stdout=stdout, stderr="",
            ))
        except Exception as exc:
            ended_at = clock.now()
            setup_results.append(CommandResult(
                template=template, command=cmd_label,
                started_at=started_at, ended_at=ended_at,
                return_code=None, stdout="", stderr="", error=str(exc),
            ))
            connection_failed = True
            break

    if connection_failed:
        try:
            shell.close()
            client.close()
        except Exception:
            pass
        server_initial = CommandResult(
            template=server_template, command=cmd_prefix + server_template,
            started_at=clock.now(), ended_at=clock.now(),
            return_code=None, stdout="", stderr="",
            error="Skipped — hub setup failed.",
        )
        return setup_results, server_initial, None

    # Start the server command on the same shell in a background thread.
    server_cmd_str = render_template(server_template, rep_site)
    cmd_label = cmd_prefix + server_cmd_str
    started_at = clock.now()
    handle = _ParamikoHandle()
    handle._shell = shell
    handle._client = client

    def _read_server() -> None:
        try:
            # Remove the per-operation timeout so the shell stays open
            # indefinitely while waiting for spoke clients.
            shell.settimeout(None)
            shell.send(server_cmd_str + "\n")
            while not getattr(shell, "closed", False):
                chunk = _shell_recv_chunk(shell)
                if chunk:
                    handle._stdout_parts.append(chunk)
                else:
                    time.sleep(0.1)
        except Exception as exc:
            handle._stderr_parts.append(str(exc))
            handle.returncode = 1
        finally:
            handle.returncode = handle.returncode if handle.returncode is not None else 0
            handle._done.set()

    threading.Thread(target=_read_server, daemon=True).start()
    server_initial = CommandResult(
        template=server_template, command=cmd_label,
        started_at=started_at, ended_at=started_at,
        return_code=None, stdout="Hub server started in global VDOM via Paramiko.", stderr="",
    )
    return setup_results, server_initial, handle


def _paramiko_spoke_session(
    site: SiteDefinition,
    templates: list[str],
    timeout: int | None,
    dry_run: bool,
    routing_intf: str | None = None,
) -> list[CommandResult]:
    """Run all spoke commands in one Paramiko shell so per-session settings persist.

    When ``routing_intf`` is given, first verify the hub WAN IP routes out via that
    client interface; if it does not, record the failed check and skip the test.
    """
    host = site.ip_address
    cmd_prefix = f"[paramiko] {_paramiko_user}@{host}: "

    if dry_run:
        now = clock.now()
        results = []
        if routing_intf:
            results.append(CommandResult(
                template=FORTIGATE_SPOKE_ROUTING_CHECK,
                command=cmd_prefix + render_template(FORTIGATE_SPOKE_ROUTING_CHECK, site),
                started_at=now, ended_at=now,
                return_code=0, stdout="dry-run: routing check not executed", stderr="",
            ))
        return results + [
            CommandResult(
                template=t, command=cmd_prefix + render_template(t, site),
                started_at=now, ended_at=now,
                return_code=0, stdout="dry-run: command not executed", stderr="",
            )
            for t in templates
        ]

    try:
        client = _paramiko_connect(host, timeout)
        shell = _paramiko_open_shell(client, timeout)
    except Exception as exc:
        now = clock.now()
        return [CommandResult(
            template=templates[0] if templates else "",
            command=cmd_prefix + "(connect)",
            started_at=now, ended_at=now,
            return_code=None, stdout="", stderr="", error=str(exc),
        )]

    results: list[CommandResult] = []
    try:
        # Pre-flight: confirm the hub WAN IP routes out via the spoke client interface.
        if routing_intf:
            rendered = render_template(FORTIGATE_SPOKE_ROUTING_CHECK, site)
            started_at = clock.now()
            try:
                shell.send(rendered + "\n")
                raw = _shell_read_until_prompt(shell, timeout=timeout or 30)
                stdout = ANSI_ESCAPE_PATTERN.sub("", raw)
                ok = routing_via_interface(stdout, routing_intf)
                results.append(CommandResult(
                    template=FORTIGATE_SPOKE_ROUTING_CHECK, command=cmd_prefix + rendered,
                    started_at=started_at, ended_at=clock.now(),
                    return_code=0 if ok else None, stdout=stdout, stderr="",
                    error=None if ok else (
                        f"Spoke route to hub {site.hub_ip} is not via the client interface "
                        f"'{routing_intf}' — skipping the traffic test."),
                ))
                if not ok:
                    return results
            except Exception as exc:
                results.append(CommandResult(
                    template=FORTIGATE_SPOKE_ROUTING_CHECK, command=cmd_prefix + rendered,
                    started_at=started_at, ended_at=clock.now(),
                    return_code=None, stdout="", stderr="", error=str(exc),
                ))
                return results

        for i, template in enumerate(templates):
            rendered = render_template(template, site)
            cmd_label = cmd_prefix + rendered
            started_at = clock.now()
            # The last command is the traffictest run — wait indefinitely for it to finish.
            cmd_timeout = None if i == len(templates) - 1 else (timeout or 30)
            try:
                shell.send(rendered + "\n")
                raw = _shell_read_until_prompt(shell, timeout=cmd_timeout)
                stdout = ANSI_ESCAPE_PATTERN.sub("", raw)
                ended_at = clock.now()
                throughput_mbps, throughput_label, retransmissions, sender_mbps, receiver_mbps = extract_throughput(stdout)
                results.append(CommandResult(
                    template=template, command=cmd_label,
                    started_at=started_at, ended_at=ended_at,
                    return_code=0, stdout=stdout, stderr="",
                    throughput_mbps=throughput_mbps,
                    throughput_label=throughput_label,
                    retransmissions=retransmissions,
                    sender_throughput_mbps=sender_mbps,
                    receiver_throughput_mbps=receiver_mbps,
                ))
            except Exception as exc:
                ended_at = clock.now()
                results.append(CommandResult(
                    template=template, command=cmd_label,
                    started_at=started_at, ended_at=ended_at,
                    return_code=None, stdout="", stderr="", error=str(exc),
                ))
                break
    finally:
        try:
            shell.close()
            client.close()
        except Exception:
            pass
    return results


def _exec_ssh(
    site: SiteDefinition,
    target: str,
    remote_template: str,
    ssh_template: str,
    timeout: int | None,
    dry_run: bool,
) -> CommandResult:
    if _use_paramiko:
        command, error = render_command_or_error(remote_template, site)
        if error is not None:
            return error
        return _paramiko_exec(target, command, remote_template, timeout, dry_run)
    command, error = build_ssh_command_or_error(site, ssh_template, target, remote_template)
    if error is not None:
        return error
    return run_rendered_command(remote_template, command, timeout=timeout, dry_run=dry_run)


def _exec_ssh_background(
    site: SiteDefinition,
    target: str,
    remote_template: str,
    ssh_template: str,
    dry_run: bool,
) -> tuple[CommandResult, subprocess.Popen[str] | _ParamikoHandle | None]:
    if _use_paramiko:
        command, error = render_command_or_error(remote_template, site)
        if error is not None:
            return error, None
        return _paramiko_start_background(target, command, remote_template, dry_run)
    command, error = build_ssh_command_or_error(site, ssh_template, target, remote_template)
    if error is not None:
        return error, None
    return start_background_command(remote_template, command, dry_run=dry_run)


def set_site_display_name(site: SiteDefinition, display_name: str) -> None:
    site.display_name = display_name
    for key in DISCOVERED_NAME_KEYS:
        site.placeholders[key] = display_name


def build_template_values(site: SiteDefinition, extra_values: dict[str, str] | None = None) -> SafeFormatDict:
    values = SafeFormatDict(site.placeholders.copy())
    values.setdefault("site_index", str(site.index))
    values.setdefault("spoke_name", site.display_name)
    values.setdefault("site_name", site.display_name)
    values.setdefault("name", site.display_name)
    values.setdefault("firewall_name", site.display_name)
    values.setdefault("hostname", site.display_name)
    values.setdefault("device_name", site.display_name)
    values.setdefault("spoke_ip", site.ip_address)
    values.setdefault("ip", site.ip_address)
    values.setdefault("hub_ip", site.hub_ip)
    values.setdefault("hub", site.hub_ip)
    values.setdefault("speed", site.speed)
    values.setdefault("expected_speed", site.speed)
    values.setdefault("speed_mbps", f"{site.speed_mbps:g}" if site.speed_mbps is not None else "")
    values.setdefault(
        "speed_with_margin_mbps",
        f"{site.speed_with_margin_mbps:g}" if site.speed_with_margin_mbps is not None else "",
    )
    values.setdefault("speed_with_margin", site.speed_with_margin_label)
    values.setdefault("bandwidth_with_margin", site.speed_with_margin_label)
    if extra_values:
        values.update(extra_values)
    return values


def render_template(
    template: str,
    site: SiteDefinition,
    extra_values: dict[str, str] | None = None,
) -> str:
    return template.format_map(build_template_values(site, extra_values))


def render_command_or_error(
    template: str,
    site: SiteDefinition,
) -> tuple[str | None, CommandResult | None]:
    try:
        command = render_template(template, site)
    except KeyError as error:
        moment = clock.now()
        return None, CommandResult(
            template=template,
            command="",
            started_at=moment,
            ended_at=moment,
            return_code=None,
            stdout="",
            stderr="",
            error=f"Missing placeholder value for '{error.args[0]}'",
        )
    return command, None


def discover_firewall_name(
    site: SiteDefinition,
    command_template: str,
    timeout: int | None,
    dry_run: bool,
) -> tuple[CommandResult, str | None]:
    if _use_paramiko:
        result = _paramiko_exec(site.ip_address, "get system status", "get system status", timeout, dry_run)
    else:
        command, error_result = render_command_or_error(command_template, site)
        if error_result is not None:
            return error_result, None
        result = run_rendered_command(command_template, command, timeout=timeout, dry_run=dry_run)

    if result.error or dry_run:
        return result, None

    discovered_name = parse_firewall_name(result.stdout)
    if discovered_name is None and result.return_code == 0:
        discovered_name = parse_firewall_name(result.stderr)
    return result, discovered_name


def run_site(
    site: SiteDefinition,
    command_templates: list[str],
    timeout: int | None,
    dry_run: bool,
    name_discovery_result: CommandResult | None = None,
) -> SiteRun:
    started_at = clock.now()
    results: list[CommandResult] = []

    for template in command_templates:
        command, error_result = render_command_or_error(template, site)
        if error_result is not None:
            results.append(error_result)
            continue
        results.append(run_rendered_command(template, command, timeout=timeout, dry_run=dry_run))

    ended_at = clock.now()
    return SiteRun(
        site=site,
        started_at=started_at,
        ended_at=ended_at,
        command_results=results,
        name_discovery_result=name_discovery_result,
    )


def run_fortigate_spoke_only(
    site: SiteDefinition,
    args: argparse.Namespace,
    name_discovery_result: CommandResult | None = None,
) -> SiteRun:
    started_at = clock.now()
    # Verify the spoke routes to the hub WAN IP via its client interface before testing.
    routing_intf = None if getattr(args, "skip_routing_check", False) else site.spoke_client_intf

    if _use_paramiko:
        results = _paramiko_spoke_session(
            site, FORTIGATE_SPOKE_COMMANDS, args.timeout, args.dry_run, routing_intf=routing_intf,
        )
    else:
        results = []
        if routing_intf and not args.dry_run:
            rchk = _exec_ssh(
                site, site.ip_address, FORTIGATE_SPOKE_ROUTING_CHECK,
                args.ssh_template, args.timeout, args.dry_run,
            )
            ok = routing_via_interface(rchk.stdout, routing_intf) if not rchk.error else False
            results.append(CommandResult(
                template=FORTIGATE_SPOKE_ROUTING_CHECK, command=rchk.command,
                started_at=rchk.started_at, ended_at=rchk.ended_at,
                return_code=rchk.return_code, stdout=rchk.stdout, stderr=rchk.stderr,
                error=rchk.error or (None if ok else
                    f"Spoke route to hub {site.hub_ip} is not via the client interface "
                    f"'{routing_intf}' — skipping the traffic test."),
            ))
        if not (results and results[-1].error):
            for remote_template in FORTIGATE_SPOKE_COMMANDS:
                results.append(
                    _exec_ssh(site, site.ip_address, remote_template, args.ssh_template, args.timeout, args.dry_run)
                )

    ended_at = clock.now()
    return SiteRun(
        site=site,
        started_at=started_at,
        ended_at=ended_at,
        command_results=results,
        name_discovery_result=name_discovery_result,
    )


def summarize(results: list[SiteRun]) -> dict[str, Any]:
    successful_sites = sum(1 for site_run in results if _compute_result(site_run)[1] == "success")
    failed_sites = len(results) - successful_sites
    sender_values = [s.max_sender_throughput_mbps for s in results if s.max_sender_throughput_mbps is not None]
    receiver_values = [s.max_receiver_throughput_mbps for s in results if s.max_receiver_throughput_mbps is not None]
    return {
        "total_sites": len(results),
        "successful_sites": successful_sites,
        "failed_sites": failed_sites,
        "peak_sender_mbps": max(sender_values) if sender_values else None,
        "peak_receiver_mbps": max(receiver_values) if receiver_values else None,
    }


def format_timestamp(value: datetime) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S")


def format_seconds(value: float) -> str:
    return f"{value:.1f}s"


def format_peak(value: float | None) -> str:
    if value is None:
        return "N/A"
    return f"{value:.2f} Mbps"


def _render_command_block(
    result: CommandResult,
    heading: str = "Template",
    status_override: tuple[str, str] | None = None,
) -> str:
    output_text = "\n".join(
        section
        for section in [
            f"STDOUT:\n{result.stdout.strip()}" if result.stdout.strip() else "",
            f"STDERR:\n{result.stderr.strip()}" if result.stderr.strip() else "",
            f"ERROR:\n{result.error}" if result.error else "",
        ]
        if section
    ) or "No output captured."
    throughput_line = (
        f'<div><strong>Detected Throughput:</strong> {html.escape(result.throughput_label)}</div>'
        if result.throughput_label is not None else ""
    )
    retr_line = (
        f'<div><strong>Retransmissions:</strong> {result.retransmissions}</div>'
        if result.retransmissions is not None else ""
    )
    return """
        <div class="command-block">
          <div><strong>{heading}:</strong> <code>{template}</code></div>
          <div><strong>Command:</strong> <code>{command}</code></div>
          <div><strong>Status:</strong> <span class="badge {status_class}">{status}</span></div>
          <div><strong>Duration:</strong> {duration}</div>
          {throughput_line}
          {retr_line}
          <pre>{output}</pre>
        </div>
        """.format(
        heading=html.escape(heading),
        template=html.escape(result.template),
        command=html.escape(result.command or "N/A"),
        status=html.escape(status_override[0] if status_override else result.status),
        status_class=html.escape(status_override[1] if status_override else result.status),
        duration=html.escape(format_seconds(result.duration_seconds)),
        throughput_line=throughput_line,
        retr_line=retr_line,
        output=html.escape(output_text),
    )


def _hub_display(site: SiteDefinition) -> str:
    if site.hub_name:
        return f"{site.hub_name} ({site.hub_ip})"
    return site.hub_ip or "N/A"


def accepted_threshold_mbps(site: SiteDefinition) -> float | None:
    """Throughput a site must reach to pass.

    Uses the device's explicit accepted_speed when set, otherwise falls back to
    DEFAULT_ACCEPT_RATIO (90%) of the configured speed.
    """
    if site.accepted_speed_mbps is not None:
        return site.accepted_speed_mbps
    if site.speed_mbps is not None:
        return round(site.speed_mbps * DEFAULT_ACCEPT_RATIO, 2)
    return None


def _compute_result(site_run: SiteRun) -> tuple[str, str]:
    """Return (result_label, css_class) for a site run based on the accepted threshold."""
    sender_mbps = site_run.max_sender_throughput_mbps
    threshold_mbps = accepted_threshold_mbps(site_run.site)
    if sender_mbps is not None and threshold_mbps is not None and sender_mbps >= threshold_mbps:
        return "Pass", "success"
    elif sender_mbps is None:
        return "Fail (not reachable)", "failed"
    else:
        return "Fail (insufficient speed)", "failed"


def _ssh_reachable(site_run: SiteRun) -> tuple[str, str]:
    """Return (label, css_class) indicating whether the SSH connection succeeded.

    Primary signal is the name-discovery step (the first SSH attempt against the
    spoke); if that wasn't run, fall back to whether any command executed without
    a connection/transport error.
    """
    ndr = site_run.name_discovery_result
    if ndr is not None:
        return ("Yes", "success") if ndr.error is None else ("No", "failed")
    if site_run.command_results:
        reachable = any(r.error is None for r in site_run.command_results)
        return ("Yes", "success") if reachable else ("No", "failed")
    return ("N/A", "skipped")


def build_html_report(
    input_path: Path,
    output_path: Path,
    results: list[SiteRun],
    command_templates: list[str],
    delay_seconds: int,
) -> str:
    summary = summarize(results)
    created_at = clock.now()

    rows_html: list[str] = []
    details_html: list[str] = []

    for site_run in results:
        result_label, result_class = _compute_result(site_run)
        sender_mbps = site_run.max_sender_throughput_mbps

        rows_html.append(
            """
            <tr data-result="{row_result}">
              <td>{index}</td>
              <td>{name}</td>
              <td>{circuit_id}</td>
              <td>{bw}</td>
              <td>{isp}</td>
              <td>{test_speed}</td>
              <td>{sender}</td>
              <td>{started}</td>
              <td><span class="badge {result_class}">{result_label}</span></td>
            </tr>
            """.format(
                index=site_run.site.index,
                name=html.escape(site_run.site.display_name),
                circuit_id=html.escape(site_run.site.circuit_id or "N/A"),
                isp=html.escape(site_run.site.isp or "N/A"),
                bw=html.escape(site_run.site.speed or "N/A"),
                test_speed=html.escape(site_run.site.speed_with_margin_label or "N/A"),
                sender=html.escape(format_peak(sender_mbps)),
                started=html.escape(format_timestamp(site_run.started_at)),
                result_label=result_label,
                result_class=result_class,
                row_result="pass" if result_class == "success" else "fail",
            )
        )

        reach_label, reach_class = _ssh_reachable(site_run)

        command_blocks: list[str] = []
        for result in site_run.command_results:
            if "traffictest run" in result.command:
                # The Status badge here reflects whether the device speed passed
                # the test (the accepted-speed threshold), not the raw command exit.
                command_blocks.append(
                    _render_command_block(result, status_override=(result_label, result_class))
                )

        sender_line = (
            f'<p><strong>Sender Throughput:</strong> {html.escape(format_peak(site_run.max_sender_throughput_mbps))}</p>'
            if site_run.max_sender_throughput_mbps is not None else ""
        )
        receiver_line = (
            f'<p><strong>Receiver Throughput:</strong> {html.escape(format_peak(site_run.max_receiver_throughput_mbps))}</p>'
            if site_run.max_receiver_throughput_mbps is not None else ""
        )
        retr_total = sum(
            r.retransmissions for r in site_run.command_results if r.retransmissions is not None
        )
        retr_line = (
            f'<p><strong>Total Retransmissions:</strong> {retr_total}</p>'
            if any(r.retransmissions is not None for r in site_run.command_results) else ""
        )
        delay_line = (
            f'<p><strong>Delay After This Site:</strong> {site_run.delayed_after_seconds}s</p>'
            if site_run.delayed_after_seconds else ""
        )

        details_html.append(
            """
            <section class="site-card">
              <h2>{name}</h2>
              <p><strong>IP:</strong> {ip}</p>
              <p><strong>Hub IP:</strong> {hub_ip}</p>
              <p><strong>Configured Speed:</strong> {speed}</p>
              <p><strong>Test Bandwidth (+15%):</strong> {test_speed}</p>
              <p><strong>Reachable:</strong> <span class="badge {status_class}">{status}</span></p>
              {sender_line}
              {receiver_line}
              {retr_line}
              <p><strong>Started:</strong> {started} &nbsp;|&nbsp; <strong>Ended:</strong> {ended} &nbsp;|&nbsp; <strong>Duration:</strong> {duration}</p>
              {delay_line}
              {commands}
            </section>
            """.format(
                name=html.escape(site_run.site.display_name),
                ip=html.escape(site_run.site.ip_address or "N/A"),
                hub_ip=html.escape(_hub_display(site_run.site)),
                speed=html.escape(site_run.site.speed or "N/A"),
                test_speed=html.escape(site_run.site.speed_with_margin_label or "N/A"),
                status=html.escape(reach_label),
                status_class=html.escape(reach_class),
                sender_line=sender_line,
                receiver_line=receiver_line,
                retr_line=retr_line,
                started=html.escape(format_timestamp(site_run.started_at)),
                ended=html.escape(format_timestamp(site_run.ended_at)),
                duration=html.escape(format_seconds(site_run.duration_seconds)),
                delay_line=delay_line,
                commands="\n".join(command_blocks),
            )
        )

    command_list = "".join(f"<li><code>{html.escape(template)}</code></li>" for template in command_templates)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>SD-WAN Traffic Test Report</title>
  <style>
    :root {{
      --bg: #f6f3ee;
      --panel: #fffdf9;
      --text: #1f2933;
      --muted: #52606d;
      --border: #d9d2c6;
      --success: #116530;
      --failed: #9b1c1c;
      --template-error: #b26b00;
      --skipped: #52606d;
      --accent: #8c3d2b;
      --accent-dark: #6e2f20;
      --accent-soft: rgba(140, 61, 43, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Tahoma, sans-serif;
      background: linear-gradient(180deg, #f4efe7 0%, #fcfaf6 100%);
      color: var(--text);
      line-height: 1.55;
    }}
    main {{
      max-width: 1200px;
      margin: 0 auto;
      padding: 36px 24px 56px;
    }}
    h1 {{ margin: 0 0 4px; font-size: 1.6rem; font-weight: 700; color: var(--accent-dark); }}
    h2 {{ margin: 0 0 14px; font-size: 1.1rem; font-weight: 600; color: var(--accent-dark); }}
    .hero {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-top: 4px solid var(--accent);
      border-radius: 12px;
      padding: 28px 32px;
      box-shadow: 0 2px 12px rgba(27,191,191,0.07);
      margin-bottom: 20px;
    }}
    .summary, .site-card {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 24px 28px;
      box-shadow: 0 2px 8px rgba(0,0,0,0.04);
      margin-bottom: 20px;
    }}
    .summary-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 14px;
      margin: 16px 0;
    }}
    .metric {{
      background: var(--accent-soft);
      border: 1px solid rgba(27,191,191,0.25);
      border-radius: 10px;
      padding: 16px;
    }}
    .metric-label {{
      color: var(--muted);
      font-size: 0.82rem;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      font-weight: 600;
    }}
    .metric-value {{
      font-size: 1.6rem;
      font-weight: 700;
      color: var(--accent-dark);
      margin-top: 4px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      margin-top: 16px;
      font-size: 0.92rem;
    }}
    th {{
      background: rgba(92,45,110,0.07);
      color: var(--accent-dark);
      font-weight: 600;
      font-size: 0.82rem;
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    th, td {{
      padding: 10px 14px;
      border-bottom: 1px solid var(--border);
      text-align: left;
      vertical-align: middle;
    }}
    tbody tr:hover {{ background: var(--accent-soft); }}
    tr:last-child td {{ border-bottom: none; }}
    code, pre {{
      font-family: "SFMono-Regular", Consolas, monospace;
      font-size: 0.88rem;
    }}
    pre {{
      white-space: pre-wrap;
      word-break: break-word;
      background: #f0f4f8;
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 14px;
      overflow-x: auto;
      margin-top: 10px;
    }}
    .command-block {{
      padding-top: 14px;
      margin-top: 14px;
      border-top: 1px solid var(--border);
    }}
    .badge {{
      display: inline-block;
      padding: 3px 12px;
      border-radius: 999px;
      font-size: 0.8rem;
      font-weight: 700;
      letter-spacing: 0.03em;
    }}
    .success {{ color: #fff; background: var(--success); }}
    .failed {{ color: #fff; background: var(--failed); }}
    .template-error {{ color: #fff; background: var(--template-error); }}
    .skipped {{ color: #fff; background: var(--skipped); }}
    .muted {{ color: var(--muted); }}
    ul {{ margin: 0; padding-left: 20px; }}
    .summary-grid .metric {{ cursor: pointer; transition: box-shadow 0.15s, transform 0.1s; }}
    .summary-grid .metric:hover {{ box-shadow: 0 0 0 2px var(--accent); transform: translateY(-1px); }}
    .summary-grid .metric.active {{ box-shadow: 0 0 0 3px var(--accent); background: var(--accent); }}
    .summary-grid .metric.active .metric-label {{ color: rgba(255,255,255,0.8); }}
    .summary-grid .metric.active .metric-value {{ color: #fff; }}
  </style>
</head>
<body>
  <main>
    <section class="hero">
      <h1>SD-WAN iPerf Traffic Test Report</h1>
      <p class="muted">Generated at {html.escape(format_timestamp(created_at))}</p>
    </section>

    <section class="summary">
      <h2>Summary</h2>
      <div class="summary-grid">
        <div class="metric" onclick="filterTable(null, this)"><div class="metric-label">Total Sites</div><div class="metric-value">{summary["total_sites"]}</div></div>
        <div class="metric" onclick="filterTable('pass', this)"><div class="metric-label">Successful Sites</div><div class="metric-value">{summary["successful_sites"]}</div></div>
        <div class="metric" onclick="filterTable('fail', this)"><div class="metric-label">Failed Sites</div><div class="metric-value">{summary["failed_sites"]}</div></div>
      </div>

      <table>
        <thead>
          <tr>
            <th>#</th>
            <th>Site name</th>
            <th>Circuit ID</th>
            <th>BW</th>
            <th>ISP</th>
            <th>Generated traffic</th>
            <th>Actual bandwidth</th>
            <th>Started</th>
            <th>Result</th>
          </tr>
        </thead>
        <tbody>
          {"".join(rows_html)}
        </tbody>
      </table>
    </section>

    {"".join(details_html)}
  </main>
  <script>
    function filterTable(result, card) {{
      var cards = document.querySelectorAll('.summary-grid .metric');
      cards.forEach(function(c) {{ c.classList.remove('active'); }});
      card.classList.add('active');
      var rows = document.querySelectorAll('tbody tr');
      rows.forEach(function(r) {{
        r.style.display = (result === null || r.dataset.result === result) ? '' : 'none';
      }});
    }}
    document.querySelector('.summary-grid .metric').classList.add('active');
  </script>
</body>
</html>
"""


def build_excel_report(results: list[SiteRun], summary: dict[str, Any], output_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print(f"openpyxl not installed — skipping Excel report. Run: {sys.executable} -m pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["#", "Site Name", "Circuit ID", "BW", "ISP", "Generated Traffic", "Actual Bandwidth", "Started", "Result"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="2D2D2D")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    pass_fill = PatternFill("solid", fgColor="116530")
    fail_fill = PatternFill("solid", fgColor="9B1C1C")
    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    white_bold = Font(color="FFFFFF", bold=True)

    for i, site_run in enumerate(results, start=2):
        result_label, result_class = _compute_result(site_run)
        ws.append([
            site_run.site.index,
            site_run.site.display_name,
            site_run.site.circuit_id or "N/A",
            site_run.site.speed or "N/A",
            site_run.site.isp or "N/A",
            site_run.site.speed_with_margin_label or "N/A",
            format_peak(site_run.max_sender_throughput_mbps),
            format_timestamp(site_run.started_at),
            result_label,
        ])
        if i % 2 == 1:
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = alt_fill
        result_cell = ws.cell(row=i, column=9)
        result_cell.fill = pass_fill if result_class == "success" else fail_fill
        result_cell.font = white_bold
        result_cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(width + 2, 12)

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.append(["Total Sites", summary["total_sites"]])
    ws2.append(["Successful Sites", summary["successful_sites"]])
    ws2.append(["Failed Sites", summary["failed_sites"]])
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    print(f"Excel report written to: {output_path}")


def build_pdf_report(results: list[SiteRun], summary: dict[str, Any], output_path: Path) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        print(f"reportlab not installed — skipping PDF report. Run: {sys.executable} -m pip install reportlab")
        return

    doc = SimpleDocTemplate(
        str(output_path), pagesize=landscape(A4),
        leftMargin=1 * cm, rightMargin=1 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("FortiGate Traffic Test Report", styles["Title"]))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        f"Total: {summary['total_sites']}  |  Pass: {summary['successful_sites']}  |  Fail: {summary['failed_sites']}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 12))

    col_headers = ["#", "Site Name", "Circuit ID", "BW", "ISP", "Generated Traffic", "Actual BW", "Started", "Result"]
    data: list[list[str]] = [col_headers]
    result_classes: list[str] = []

    for site_run in results:
        result_label, result_class = _compute_result(site_run)
        result_classes.append(result_class)
        data.append([
            str(site_run.site.index),
            site_run.site.display_name,
            site_run.site.circuit_id or "N/A",
            site_run.site.speed or "N/A",
            site_run.site.isp or "N/A",
            site_run.site.speed_with_margin_label or "N/A",
            format_peak(site_run.max_sender_throughput_mbps),
            format_timestamp(site_run.started_at),
            result_label,
        ])

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D2D2D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]
    for i, result_class in enumerate(result_classes, start=1):
        bg = colors.HexColor("#116530") if result_class == "success" else colors.HexColor("#9B1C1C")
        style_cmds += [
            ("BACKGROUND", (8, i), (8, i), bg),
            ("TEXTCOLOR", (8, i), (8, i), colors.white),
            ("FONTNAME", (8, i), (8, i), "Helvetica-Bold"),
        ]

    table = Table(data, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)
    doc.build(elements)
    print(f"PDF report written to: {output_path}")


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run FortiGate hub/spoke traffictest commands sequentially from CSV/XLSX input and generate an HTML report."
    )
    parser.add_argument("--input", default="devices.csv", help="Path to a CSV or XLSX file containing spoke data (default: devices.csv).")
    parser.add_argument("--sheet", help="Worksheet name to read when the input file is XLSX.")
    parser.add_argument(
        "--command",
        action="append",
        help=(
            "Optional custom command template to run for each site. If omitted, the built-in "
            "FortiGate diagnose traffictest hub/spoke flow is used. Useful placeholders include "
            "{firewall_name}, {spoke_name}, {spoke_ip}, {speed}, "
            "{speed_with_margin}, {speed_with_margin_mbps}, and {hub_ip}."
        ),
    )
    parser.add_argument(
        "--command-file",
        help="Text file containing one command template per line. Blank lines and lines starting with # are ignored.",
    )
    parser.add_argument(
        "--firewall-name-command",
        default='ssh -o BatchMode=yes -o ConnectTimeout=10 -o StrictHostKeyChecking=accept-new {spoke_ip} "get system status"',
        help=(
            "SSH command template used to discover the firewall name before each test. "
            "The output may contain a line like 'Hostname: FW-01' or just the hostname. "
            "Default: %(default)s"
        ),
    )
    parser.add_argument(
        "--firewall-name-timeout",
        type=int,
        default=30,
        help="Timeout in seconds for firewall name discovery. Default: 30.",
    )
    parser.add_argument(
        "--sshuser",
        help="SSH username to prepend to every target (e.g. admin).",
    )
    parser.add_argument(
        "--sshpw",
        nargs="?",
        const=True,
        default=None,
        metavar="PASSWORD",
        help=(
            "SSH password. Supply it directly as a value, or pass the flag with no value "
            "to be prompted interactively with hidden characters. Uses sshpass on Linux/macOS "
            "and plink on Windows (see --plink)."
        ),
    )
    parser.add_argument(
        "--plink",
        nargs="?",
        const="plink",
        default=None,
        metavar="PATH",
        help=(
            "Use PuTTY plink instead of ssh/sshpass. Optionally supply the full path to "
            "plink.exe; omit the value to use 'plink' from PATH. When combined with "
            "--sshpw, passes -pw to plink instead of using sshpass. "
            "Note: host keys must be pre-accepted in PuTTY's registry cache before "
            "running in batch mode."
        ),
    )
    parser.add_argument(
        "--paramiko",
        action="store_true",
        default=True,
        help=(
            "Use the Paramiko Python library for SSH instead of external binaries. "
            "Requires: pip install paramiko. Enabled by default."
        ),
    )
    parser.add_argument(
        "--hub-ip",
        help="Hub firewall IP address used by spokes for iperf3. If omitted, each row must provide a hub_ip column.",
    )
    parser.add_argument(
        "--hub-mgmt-ip",
        help=(
            "Hub management IP address used for SSH (setup and server commands). "
            "Falls back to hub_ip when omitted. Can also be set per row with a hub_mgmt_ip column."
        ),
    )
    parser.add_argument(
        "--ssh-template",
        default=DEFAULT_SSH_TEMPLATE,
        help=(
            "SSH wrapper template for built-in FortiGate traffictest commands. "
            "Available placeholders: {target}, {remote_command}, plus site placeholders. "
            "Default: %(default)s"
        ),
    )
    parser.add_argument(
        "--hub-server-intf",
        default=DEFAULT_HUB_SERVER_INTF,
        help=(
            "Fallback hub interface for 'diagnose traffictest server-intf' when the input row has no "
            f"server_intf/hub_server_intf column. Default: {DEFAULT_HUB_SERVER_INTF}."
        ),
    )
    parser.add_argument(
        "--spoke-client-intf",
        default=DEFAULT_SPOKE_CLIENT_INTF,
        help=(
            "Fallback spoke interface for 'diagnose traffictest client-intf' when the input row has no "
            f"client_intf/spoke_client_intf column. Default: {DEFAULT_SPOKE_CLIENT_INTF}."
        ),
    )
    parser.add_argument(
        "--traffictest-port",
        default=DEFAULT_TRAFFICTEST_PORT,
        help=(
            "Fallback FortiGate traffictest port when the input row has no "
            f"traffictest_port/traffic_port column. Default: {DEFAULT_TRAFFICTEST_PORT}."
        ),
    )
    parser.add_argument(
        "--traffictest-duration",
        type=int,
        default=None,
        help=(
            "Run-wide test duration in seconds, appended as '-t <n>' to the spoke "
            "'diagnose traffictest run' command. A per-device traffictest_duration "
            "column overrides this. If neither is set, no -t is sent and FortiGate "
            f"uses its built-in {DEFAULT_TRAFFICTEST_DURATION_SECONDS}s default."
        ),
    )
    parser.add_argument(
        "--hub-server-start-delay",
        type=float,
        default=DEFAULT_HUB_SERVER_START_DELAY_SECONDS,
        help=(
            "Seconds to wait after starting the hub traffictest server before running spoke commands. "
            f"Default: {DEFAULT_HUB_SERVER_START_DELAY_SECONDS}."
        ),
    )
    parser.add_argument(
        "--delay-seconds",
        type=int,
        default=DEFAULT_DELAY_SECONDS,
        help=f"Delay between sites in seconds. Default: {DEFAULT_DELAY_SECONDS}.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=None,
        help="Optional timeout per command in seconds.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="HTML report output path. Default: traffic_test_report_YYYYMMDD_HHMMSS.html",
    )
    parser.add_argument(
        "--skip-hub-setup",
        action="store_true",
        help=(
            "Skip all hub SSH commands (config global, server-intf, port, run -s). "
            "Use this when you have already started the hub traffictest server manually. "
            "The script will run only the spoke-side test commands."
        ),
    )
    parser.add_argument(
        "--skip-speedtest-check",
        action="store_true",
        help=(
            "Skip the pre-flight check that verifies the hub server interface permits "
            "'speed-test' in its allowaccess before starting the traffictest server."
        ),
    )
    parser.add_argument(
        "--skip-routing-check",
        action="store_true",
        help=(
            "Skip the pre-flight routing checks: on the hub that each spoke IP routes "
            "via the server interface (failed spokes are skipped before the server "
            "starts), and on the spoke that the hub WAN IP routes via the client interface."
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Render commands and report without executing the traffic tests.",
    )
    return parser


def _make_speedometer_icon_b64() -> str:
    """Generate a 32×32 speedometer PNG icon and return it as a base64 string (stdlib only)."""
    import base64
    import math
    import struct
    import zlib

    W = H = 32
    px: list[list[tuple[int, int, int, int]]] = [[(0, 0, 0, 0)] * W for _ in range(H)]

    def setp(x: int, y: int, r: int, g: int, b: int, a: int = 255) -> None:
        if 0 <= x < W and 0 <= y < H:
            px[y][x] = (r, g, b, a)

    cx, cy = 16, 17

    # Dark navy background circle
    for y in range(H):
        for x in range(W):
            if (x - cx) ** 2 + (y - cy) ** 2 <= 14 ** 2:
                setp(x, y, 20, 40, 75)

    # Thin lighter-blue outer ring
    for y in range(H):
        for x in range(W):
            d = ((x - cx) ** 2 + (y - cy) ** 2) ** 0.5
            if 13.2 <= d <= 14.5:
                setp(x, y, 70, 110, 170)

    # bearing_to_screen: 0° = up (12 o'clock), clockwise
    def bxy(bearing: float, radius: float) -> tuple[float, float]:
        rad = math.radians(bearing)
        return cx + radius * math.sin(rad), cy - radius * math.cos(rad)

    START = 240   # bearing: lower-left (8 o'clock) = 0 speed
    SPAN  = 240   # degrees clockwise to lower-right (4 o'clock) = full speed

    # Light-blue arc
    for step in range(SPAN + 1):
        bx, by = bxy((START + step) % 360, 10.5)
        ix, iy = int(round(bx)), int(round(by))
        for dy in range(-1, 2):
            for dx in range(-1, 2):
                nx, ny = ix + dx, iy + dy
                d = ((nx - cx) ** 2 + (ny - cy) ** 2) ** 0.5
                if 9.5 <= d <= 12.0:
                    setp(nx, ny, 150, 200, 240)

    # 5 white tick marks at 0 / 25 / 50 / 75 / 100 %
    for pct in (0, 25, 50, 75, 100):
        brg = (START + pct * SPAN // 100) % 360
        ox, oy = bxy(brg, 8.0)
        ix2, iy2 = bxy(brg, 11.5)
        for s in range(9):
            t = s / 8
            setp(int(round(ox + t * (ix2 - ox))), int(round(oy + t * (iy2 - oy))), 255, 255, 255)

    # Orange needle at 75 % (2 o'clock = fast)
    needle_brg = (START + int(0.75 * SPAN)) % 360
    nx2, ny2 = bxy(needle_brg, 8.5)
    for s in range(13):
        t = s / 12
        lx = int(round(cx + t * (nx2 - cx)))
        ly = int(round(cy + t * (ny2 - cy)))
        setp(lx, ly, 255, 140, 0)
        setp(lx, ly - 1, 255, 140, 0)  # one extra pixel for thickness

    # Orange center dot
    for dy in range(-2, 3):
        for dx in range(-2, 3):
            if dx * dx + dy * dy <= 4:
                setp(cx + dx, cy + dy, 255, 160, 0)

    # Encode as PNG (RGBA, 8-bit, no interlace)
    def png_chunk(tag: bytes, data: bytes) -> bytes:
        payload = tag + data
        return struct.pack(">I", len(data)) + payload + struct.pack(">I", zlib.crc32(payload) & 0xFFFFFFFF)

    ihdr = struct.pack(">IIBBBBB", W, H, 8, 6, 0, 0, 0)
    raw = b"".join(b"\x00" + bytes([c for rgba in row for c in rgba]) for row in px)
    png = (
        b"\x89PNG\r\n\x1a\n"
        + png_chunk(b"IHDR", ihdr)
        + png_chunk(b"IDAT", zlib.compress(raw, 9))
        + png_chunk(b"IEND", b"")
    )
    return base64.b64encode(png).decode()


_ICON_B64: str | None = None
_gui_msg_queue: "queue.Queue[Any] | None" = None


def _apply_window_icon(root: Any) -> None:
    global _ICON_B64
    try:
        import tkinter as tk
        if _ICON_B64 is None:
            _ICON_B64 = _make_speedometer_icon_b64()
        img = tk.PhotoImage(data=_ICON_B64)
        root.iconphoto(True, img)
        root._icon_img = img  # keep a reference so GC doesn't collect it
    except Exception:
        pass


class _TeeWriter:
    """Forwards writes to both the original stream and a queue (line-buffered for the GUI)."""

    def __init__(self, original: Any, q: "queue.Queue[str | None]") -> None:
        self._original = original
        self._q = q
        self._buf = ""

    def write(self, text: str) -> int:
        self._original.write(text)
        self._buf += text
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            self._q.put(line)
        return len(text)

    def flush(self) -> None:
        self._original.flush()

    def fileno(self) -> int:
        return self._original.fileno()


def _show_progress_window(target_fn: "Any") -> int:
    """Run target_fn() in a background thread, streaming its stdout into a GUI log window."""
    try:
        import tkinter as tk
        from tkinter import messagebox

        global _gui_msg_queue
        msg_q: queue.Queue[Any] = queue.Queue()
        _gui_msg_queue = msg_q
        result_holder: list[int] = [0]

        root = tk.Tk()
        root.title("FortiGate Traffic Test Runner — Progress")
        root.geometry("800x500")
        root.configure(bg="#1e1e1e")
        root.attributes("-topmost", True)
        _apply_window_icon(root)

        tk.Label(
            root, text="FortiGate Traffic Test Runner",
            bg="#1e1e1e", fg="#ffffff", font=("Segoe UI", 11, "bold"),
        ).pack(pady=(12, 4))

        frame = tk.Frame(root, bg="#1e1e1e")
        frame.pack(fill="both", expand=True, padx=10, pady=(0, 5))

        sb = tk.Scrollbar(frame)
        sb.pack(side="right", fill="y")

        text_widget = tk.Text(
            frame, wrap="word", state="disabled",
            bg="#1e1e1e", fg="#d4d4d4",
            font=("Courier New", 9), yscrollcommand=sb.set,
            borderwidth=0, relief="flat",
        )
        text_widget.pack(side="left", fill="both", expand=True)
        sb.config(command=text_widget.yview)

        close_btn = tk.Button(
            root, text="Close", state="disabled", command=root.destroy, width=12,
        )
        close_btn.pack(pady=(0, 12))

        original_stdout = sys.stdout
        sys.stdout = _TeeWriter(original_stdout, msg_q)

        def _worker() -> None:
            try:
                result_holder[0] = target_fn()
            except SystemExit as exc:
                result_holder[0] = exc.code if isinstance(exc.code, int) else 1
                msg_q.put(f"Exited with code {result_holder[0]}.")
            except Exception as exc:
                result_holder[0] = 1
                msg_q.put(f"Unexpected error: {exc}")
            finally:
                sys.stdout = original_stdout
                msg_q.put(None)

        def _poll() -> None:
            global _gui_msg_queue
            try:
                while True:
                    msg = msg_q.get_nowait()
                    if msg is None:
                        _gui_msg_queue = None
                        close_btn.config(state="normal")
                        root.attributes("-topmost", False)
                        return
                    if isinstance(msg, tuple) and msg[0] == "error_dialog":
                        messagebox.showerror(msg[1], msg[2], parent=root)
                    else:
                        text_widget.config(state="normal")
                        text_widget.insert("end", msg + "\n")
                        text_widget.see("end")
                        text_widget.config(state="disabled")
            except queue.Empty:
                pass
            root.after(100, _poll)

        threading.Thread(target=_worker, daemon=True).start()
        root.after(100, _poll)
        root.mainloop()
        return result_holder[0]

    except Exception:
        return target_fn()


def _show_inputs_dialog(default_file: str) -> tuple[str, str, str] | None:
    """Show a GUI form for file path, username, and password. Returns (file, user, pw) or None on failure."""
    try:
        import tkinter as tk
        from tkinter import filedialog

        saved_user, saved_pw = load_credentials()
        result: list[tuple[str, str, str] | None] = [None]

        root = tk.Tk()
        root.title("FortiGate Traffic Test Runner")
        root.resizable(False, False)
        root.attributes("-topmost", True)
        _apply_window_icon(root)

        pad: dict = {"padx": 10, "pady": 6}

        tk.Label(root, text="Input file (CSV or XLSX):").grid(row=0, column=0, sticky="w", **pad)
        file_var = tk.StringVar(value=default_file)
        tk.Entry(root, textvariable=file_var, width=38).grid(row=0, column=1, **pad)

        def browse() -> None:
            path = filedialog.askopenfilename(
                title="Select input file",
                filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
            )
            if path:
                file_var.set(path)

        tk.Button(root, text="Browse…", command=browse).grid(row=0, column=2, padx=(0, 10), pady=6)

        tk.Label(root, text="SSH username:").grid(row=1, column=0, sticky="w", **pad)
        user_var = tk.StringVar(value=saved_user)
        tk.Entry(root, textvariable=user_var, width=38).grid(row=1, column=1, columnspan=2, sticky="w", **pad)

        tk.Label(root, text="SSH password:").grid(row=2, column=0, sticky="w", **pad)
        pw_var = tk.StringVar(value=saved_pw)
        tk.Entry(root, textvariable=pw_var, show="*", width=38).grid(row=2, column=1, columnspan=2, sticky="w", **pad)

        save_var = tk.BooleanVar(value=bool(saved_user or saved_pw))
        tk.Checkbutton(root, text="Save credentials", variable=save_var).grid(
            row=3, column=0, columnspan=3, sticky="w", padx=10, pady=(4, 0)
        )

        def on_ok() -> None:
            user = user_var.get().strip()
            pw = pw_var.get()
            if save_var.get():
                save_credentials(user, pw)
            elif _CREDENTIALS_FILE.exists():
                _CREDENTIALS_FILE.unlink()
            result[0] = (file_var.get().strip().strip('"').strip("'"), user, pw)
            root.destroy()

        def on_cancel() -> None:
            root.destroy()

        btn_frame = tk.Frame(root)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=10)
        tk.Button(btn_frame, text="OK", width=10, command=on_ok).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Cancel", width=10, command=on_cancel).pack(side="left", padx=5)

        root.bind("<Return>", lambda _e: on_ok())
        root.bind("<Escape>", lambda _e: on_cancel())

        root.update_idletasks()
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        w, h = root.winfo_reqwidth(), root.winfo_reqheight()
        root.geometry(f"+{(sw - w) // 2}+{(sh - h) // 2}")

        root.mainloop()
        return result[0]
    except Exception:
        return None


_CREDENTIALS_FILE = Path(__file__).parent / "credentials.json"


def _secret_store():
    # Lazy import so the CLI without an SSH save/load run doesn't pay the
    # cryptography import cost or trigger key creation.
    import secret_store
    return secret_store


def load_credentials() -> tuple[str, str]:
    """Return (username, password) from credentials.json, or ('', '') if not found."""
    try:
        data = json.loads(_CREDENTIALS_FILE.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return "", ""
    username = data.get("username", "")
    password = data.get("password", "")
    if password:
        try:
            password = _secret_store().decrypt(password)
        except Exception:
            # If the key is unavailable or the token is broken, fall back to
            # the literal value (covers a freshly-encrypted file the user
            # then moved to a machine without the key).
            password = ""
    return username, password


def save_credentials(username: str, password: str) -> None:
    """Write username and password to credentials.json — password encrypted at rest."""
    encrypted_pw = _secret_store().encrypt(password) if password else ""
    _CREDENTIALS_FILE.write_text(
        json.dumps({"username": username, "password": encrypted_pw}, indent=2),
        encoding="utf-8",
    )


def prompt_interactive_inputs(args: argparse.Namespace) -> None:
    print("=" * 60)
    print("FortiGate Traffic Test Runner — Interactive Mode")
    print("=" * 60)

    dialog_result = _show_inputs_dialog(args.input)
    if dialog_result is not None:
        file_path, username, password = dialog_result
        if file_path:
            args.input = file_path
        if username:
            args.sshuser = username
        if password:
            args.sshpw = password
    else:
        # Fallback to console if tkinter is unavailable
        raw = input(f"Input file path (CSV or XLSX) [{args.input}]: ").strip().strip('"').strip("'")
        if raw:
            args.input = raw
        saved_user, saved_pw = load_credentials()
        username = input(f"SSH username [{saved_user or 'blank to skip'}]: ").strip()
        if not username and saved_user:
            username = saved_user
        if username:
            args.sshuser = username
        password = getpass.getpass("SSH password (leave blank to use saved): ")
        if not password and saved_pw:
            password = saved_pw
        if password:
            args.sshpw = password


def _run_tests(args: argparse.Namespace, parser: argparse.ArgumentParser) -> int:
    # Apply saved credentials as defaults when no CLI override is provided.
    saved_user, saved_pw = load_credentials()
    if not args.sshuser and saved_user:
        args.sshuser = saved_user
    if args.sshpw is None and saved_pw:
        args.sshpw = saved_pw

    input_path = Path(args.input).expanduser().resolve()
    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        reports_dir = Path("Reports").resolve()
        reports_dir.mkdir(exist_ok=True)
        output_path = reports_dir / f"traffic_test_report_{clock.now().strftime('%Y%m%d_%H%M%S')}.html"

    if not input_path.exists():
        parser.error(f"Input file not found: {input_path}")

    if args.delay_seconds < 0:
        parser.error("--delay-seconds must be 0 or greater.")
    if args.firewall_name_timeout < 1:
        parser.error("--firewall-name-timeout must be 1 or greater.")
    if args.hub_server_start_delay < 0:
        parser.error("--hub-server-start-delay must be 0 or greater.")

    if args.paramiko or args.plink or args.sshuser or args.sshpw is not None:
        password: str | None = None
        if args.sshpw is not None:
            password = getpass.getpass("SSH password: ") if args.sshpw is True else args.sshpw

        if args.paramiko:
            if not _PARAMIKO_OK:
                detail = f" (import error: {_PARAMIKO_IMPORT_ERROR})" if _PARAMIKO_IMPORT_ERROR else ""
                parser.error(
                    f"--paramiko requires Paramiko. Install it with: pip install paramiko{detail}\n"
                    f"  Running interpreter: {sys.executable}"
                )
            global _use_paramiko, _paramiko_user, _paramiko_pass
            _use_paramiko = True
            _paramiko_user = args.sshuser or ""
            _paramiko_pass = password or ""
        else:
            user_at = f"{args.sshuser}@" if args.sshuser else ""
            if args.plink:
                pw_part = f" -pw {shlex.quote(password)}" if password else ""
                ssh_base = f"{shlex.quote(args.plink)} -batch -ssh{pw_part}"
            elif password:
                ssh_base = f"sshpass -p {shlex.quote(password)} ssh -o ConnectTimeout=10 -o StrictHostKeyChecking=accept-new"
            else:
                ssh_base = "ssh -o BatchMode=yes -o ConnectTimeout=10 -o StrictHostKeyChecking=accept-new"
            args.ssh_template = f'{ssh_base} {user_at}{{target}} "{{remote_command}}"'
            args.firewall_name_command = f'{ssh_base} {user_at}{{spoke_ip}} "get system status"'

    rows = load_rows(input_path, args.sheet)
    if not rows:
        parser.error("No spoke rows were found in the input file.")

    sites = build_sites(rows)
    for site in sites:
        if args.hub_ip:
            site.hub_ip = args.hub_ip
            site.placeholders["hub_ip"] = args.hub_ip
            site.placeholders["hub"] = args.hub_ip
        if args.hub_mgmt_ip:
            site.hub_mgmt_ip = args.hub_mgmt_ip
            site.placeholders["hub_mgmt_ip"] = args.hub_mgmt_ip
        site.hub_server_intf = site.hub_server_intf or args.hub_server_intf
        site.spoke_client_intf = site.spoke_client_intf or args.spoke_client_intf
        site.traffictest_port = site.traffictest_port or str(args.traffictest_port)
        site.placeholders["hub_server_intf"] = site.hub_server_intf
        site.placeholders["spoke_client_intf"] = site.spoke_client_intf
        site.placeholders["traffictest_port"] = site.traffictest_port
        site.placeholders["traffic_port"] = site.traffictest_port
        # Test duration: per-device value wins, else the run-wide --traffictest-duration
        # if one was supplied. A blank/invalid value means no -t (FortiGate default 10s).
        site.traffictest_duration = _sanitize_duration(site.traffictest_duration) or (
            _sanitize_duration(args.traffictest_duration) if args.traffictest_duration else ""
        )
        site.placeholders["traffictest_duration"] = site.traffictest_duration
        site.placeholders["duration_flag"] = (
            f" -t {site.traffictest_duration}" if site.traffictest_duration else ""
        )

    command_templates = load_command_templates(args)
    use_builtin_traffictest = not command_templates
    active_command_templates = fortigate_traffictest_templates() if use_builtin_traffictest else command_templates

    if use_builtin_traffictest:
        missing_hub_rows = [str(site.index) for site in sites if not site.hub_ip]
        if missing_hub_rows:
            parser.error(
                "The built-in FortiGate traffictest flow requires --hub-ip or a hub_ip column. "
                f"Missing hub IP for row(s): {', '.join(missing_hub_rows)}"
            )
        missing_speed_rows = [str(site.index) for site in sites if not site.speed_with_margin_label]
        if missing_speed_rows:
            parser.error(
                "The built-in FortiGate traffictest flow requires a speed value for every row. "
                f"Missing/invalid speed for row(s): {', '.join(missing_speed_rows)}"
            )

    available_placeholders = set(sites[0].placeholders)
    available_placeholders.update(
        {
            "site_index",
            "spoke_name",
            "site_name",
            "name",
            "firewall_name",
            "hostname",
            "device_name",
            "spoke_ip",
            "ip",
            "hub_ip",
            "hub",
            "speed",
            "expected_speed",
            "speed_mbps",
            "speed_with_margin_mbps",
            "speed_with_margin",
            "bandwidth_with_margin",
            "hub_server_intf",
            "spoke_client_intf",
            "traffictest_port",
            "traffic_port",
            "traffictest_duration",
            "duration_flag",
        }
    )
    validate_template_fields(active_command_templates, available_placeholders)
    validate_template_fields([args.firewall_name_command], available_placeholders)
    if use_builtin_traffictest and not _use_paramiko:
        validate_template_fields(
            [args.ssh_template],
            available_placeholders | {"target", "remote_command"},
        )

    runs: list[SiteRun] = []
    hub_contexts: dict[str, dict] = {}
    total_sites = len(sites)

    if use_builtin_traffictest:
        # Collect unique hub IPs in the order they first appear, with a representative site each.
        seen_hubs: dict[str, SiteDefinition] = {}
        for site in sites:
            if site.hub_ip and site.hub_ip not in seen_hubs:
                seen_hubs[site.hub_ip] = site

        # Group spokes by hub IP into per-hub queues, preserving the original row order.
        hub_queues: dict[str, list[SiteDefinition]] = {hub_ip: [] for hub_ip in seen_hubs}
        for site in sites:
            hub_queues[site.hub_ip].append(site)

        hub_contexts_lock = threading.Lock()

        def _setup_one_hub(hub_ip: str, rep_site: SiteDefinition) -> None:
            ssh_target = rep_site.hub_mgmt_ip or hub_ip
            # Per-spoke routing verdicts for this hub ({spoke_ip: ok}). Failed spokes
            # are skipped in their queue; this is kept out of the hub-level
            # connection_failed so one bad route doesn't fail the whole hub.
            routing_verdicts: dict[str, bool] = {}
            routing_results: list[CommandResult] = []
            spokes_here = hub_queues.get(hub_ip, [])

            if _use_paramiko:
                if not args.skip_routing_check:
                    routing_results, routing_verdicts = _paramiko_hub_routing_check(
                        ssh_target, spokes_here, rep_site.hub_server_intf,
                        args.timeout, args.dry_run,
                    )
                # One shell session: config global once, then all commands.
                setup_results, server_initial, server_process = _paramiko_hub_session(
                    ssh_target, rep_site,
                    FORTIGATE_HUB_SETUP_COMMANDS, FORTIGATE_HUB_SERVER_COMMAND,
                    args.timeout, args.dry_run,
                    check_speedtest=not args.skip_speedtest_check,
                )
                connection_failed = any(r.error for r in setup_results) or (
                    server_initial is not None and server_initial.error == "Skipped — hub setup failed."
                )
            else:
                # Subprocess: each SSH call is a fresh session, so prepend
                # "config global" transparently to the remote command.
                setup_results = []
                connection_failed = False

                # Pre-flight: per-spoke routing via the hub server interface.
                if not args.skip_routing_check and not args.dry_run:
                    for s in spokes_here:
                        spoke_ip = s.ip_address or ""
                        rchk = _exec_ssh(
                            rep_site, ssh_target,
                            f"get router info routing-table details {spoke_ip}",
                            args.ssh_template, args.timeout, args.dry_run,
                        )
                        ok = routing_via_interface(rchk.stdout, rep_site.hub_server_intf) if not rchk.error else False
                        routing_verdicts[spoke_ip] = ok
                        routing_results.append(CommandResult(
                            template=FORTIGATE_HUB_ROUTING_CHECK, command=rchk.command,
                            started_at=rchk.started_at, ended_at=rchk.ended_at,
                            return_code=rchk.return_code, stdout=rchk.stdout, stderr=rchk.stderr,
                            error=rchk.error or (None if ok else
                                f"Spoke {spoke_ip} does not route via the hub server interface "
                                f"'{rep_site.hub_server_intf}' — skipping this spoke."),
                        ))

                # Pre-flight: verify speed-test is permitted on the hub server interface.
                if not args.skip_speedtest_check and not args.dry_run:
                    intf = render_template("{hub_server_intf}", rep_site)
                    chk = _exec_ssh(
                        rep_site, ssh_target,
                        "config global\n" + FORTIGATE_HUB_ALLOWACCESS_CHECK,
                        args.ssh_template, args.timeout, args.dry_run,
                    )
                    allowed = speedtest_allowed(chk.stdout) if not chk.error else None
                    if chk.error or allowed is False:
                        err = chk.error or (
                            f"Hub interface '{intf}' does not permit speed-test in allowaccess. "
                            f"Enable it on the hub:  config system interface / edit {intf} / "
                            f"append allowaccess speed-test / end")
                        setup_results.append(CommandResult(
                            template=FORTIGATE_HUB_ALLOWACCESS_CHECK, command=chk.command,
                            started_at=chk.started_at, ended_at=chk.ended_at,
                            return_code=chk.return_code, stdout=chk.stdout, stderr=chk.stderr,
                            error=err,
                        ))
                        connection_failed = True
                    else:
                        note = (f"speed-test is permitted on '{intf}'." if allowed
                                else f"Could not read allowaccess for '{intf}'; proceeding without the speed-test gate.")
                        setup_results.append(CommandResult(
                            template=FORTIGATE_HUB_ALLOWACCESS_CHECK, command=chk.command,
                            started_at=chk.started_at, ended_at=chk.ended_at,
                            return_code=0, stdout=note + "\n\n" + (chk.stdout or ""), stderr="",
                        ))

                for template in (FORTIGATE_HUB_SETUP_COMMANDS if not connection_failed else []):
                    result = _exec_ssh(
                        rep_site, ssh_target,
                        "config global\n" + template,
                        args.ssh_template, args.timeout, args.dry_run,
                    )
                    # Report uses the clean template, not the prefixed one.
                    result = CommandResult(
                        template=template, command=result.command,
                        started_at=result.started_at, ended_at=result.ended_at,
                        return_code=result.return_code,
                        stdout=result.stdout, stderr=result.stderr,
                        error=result.error,
                    )
                    setup_results.append(result)
                    if result.error:
                        connection_failed = True
                        break

                server_initial = None
                server_process = None
                if not connection_failed:
                    server_initial, server_process = _exec_ssh_background(
                        rep_site, ssh_target,
                        "config global\n" + FORTIGATE_HUB_SERVER_COMMAND,
                        args.ssh_template, args.dry_run,
                    )
                    if server_initial is not None:
                        server_initial = CommandResult(
                            template=FORTIGATE_HUB_SERVER_COMMAND,
                            command=server_initial.command,
                            started_at=server_initial.started_at, ended_at=server_initial.ended_at,
                            return_code=server_initial.return_code,
                            stdout=server_initial.stdout, stderr=server_initial.stderr,
                            error=server_initial.error,
                        )

            hub_discovered_name = ""
            if not connection_failed and not args.dry_run:
                hub_site = SiteDefinition(
                    index=0, raw={},
                    placeholders={"spoke_ip": ssh_target},
                    display_name=ssh_target,
                    ip_address=ssh_target,
                    hub_ip=hub_ip,
                    speed="", speed_mbps=None,
                    speed_with_margin_mbps=None,
                    speed_with_margin_label="",
                )
                _, hub_discovered_name = discover_firewall_name(
                    hub_site, args.firewall_name_command,
                    timeout=args.firewall_name_timeout, dry_run=args.dry_run,
                )
                hub_discovered_name = hub_discovered_name or ""

            with hub_contexts_lock:
                hub_contexts[hub_ip] = {
                    "ssh_target": ssh_target,
                    "setup_results": setup_results,
                    "server_initial": server_initial,
                    "server_process": server_process,
                    "failed": connection_failed,
                    "hub_name": hub_discovered_name,
                    "routing": routing_verdicts,
                    "routing_results": routing_results,
                }

        if not args.skip_hub_setup:
            print(f"Starting traffictest server on {len(seen_hubs)} hub(s) in parallel...", flush=True)
            setup_threads = [
                threading.Thread(target=_setup_one_hub, args=(hub_ip, rep_site), daemon=True)
                for hub_ip, rep_site in seen_hubs.items()
            ]
            for t in setup_threads:
                t.start()
            for t in setup_threads:
                t.join()

            # Propagate discovered hub names to each spoke site.
            for site in sites:
                site.hub_name = hub_contexts.get(site.hub_ip, {}).get("hub_name", "")

            # Detect any hub server that exited before the delay (error condition).
            for _, ctx in hub_contexts.items():
                proc = ctx["server_process"]
                if proc is not None and proc.poll() is not None:
                    ctx["server_initial"] = finalize_background_command(
                        ctx["server_initial"], proc, stop_if_running=False
                    )
                    ctx["server_process"] = None

            failed_hubs = [ip for ip, ctx in hub_contexts.items() if ctx.get("failed")]
            if failed_hubs:
                error_lines: list[str] = []
                for ip in failed_hubs:
                    ctx_results = hub_contexts[ip].get("setup_results", [])
                    err = next((r.error for r in ctx_results if r.error), None) or "unknown error"
                    line = f"Hub {ip}: {err}"
                    error_lines.append(line)
                    print(f"  {line}", flush=True)
                if len(failed_hubs) == len(hub_contexts):
                    msg = "Could not connect to any hub:\n\n" + "\n".join(error_lines)
                    print("All hubs failed to connect — aborting.", flush=True)
                    if _gui_msg_queue is not None:
                        _gui_msg_queue.put(("error_dialog", "Hub Connection Failed", msg))
                    now = clock.now()
                    runs = []
                    for site in sites:
                        runs.append(SiteRun(site=site, started_at=now, ended_at=now))
                    report_html = build_html_report(
                        input_path=input_path, output_path=output_path,
                        results=runs, command_templates=active_command_templates,
                        delay_seconds=args.delay_seconds,
                    )
                    output_path.write_text(report_html, encoding="utf-8")
                    print(f"Report written to: {output_path}")
                    return 1
                partial_msg = "\n".join(error_lines)
                if _gui_msg_queue is not None:
                    _gui_msg_queue.put((
                        "error_dialog", "Some Hubs Failed",
                        f"{len(failed_hubs)} of {len(hub_contexts)} hub(s) failed — "
                        f"their spokes will be skipped.\n\n{partial_msg}",
                    ))
                print(
                    f"{len(failed_hubs)} of {len(hub_contexts)} hub(s) failed — "
                    f"their spokes will be skipped. Waiting {args.hub_server_start_delay:.0f}s "
                    f"for remaining hub(s)...",
                    flush=True,
                )
            else:
                print(f"Waiting {args.hub_server_start_delay:.0f}s for all hub servers to be ready...", flush=True)
            time.sleep(args.hub_server_start_delay)
        else:
            print("Hub setup skipped — assuming hub traffictest server is already running.", flush=True)

        # Run each hub's queue in its own thread (hubs run in parallel; spokes per hub run
        # sequentially so only one spoke at a time is active against each hub server).
        all_runs: list[SiteRun] = []
        all_runs_lock = threading.Lock()
        print_lock = threading.Lock()
        # Shared, global completion counter across all hub queues. Emitted at
        # column 0 as "[done/total] ..." so the webapp progress parser (which is
        # anchored at the start of the line) can follow overall progress.
        progress_lock = threading.Lock()
        progress = {"done": 0}

        def _emit_progress(message: str) -> None:
            with progress_lock:
                progress["done"] += 1
                done = progress["done"]
            with print_lock:
                print(f"[{done}/{total_sites}] {message}", flush=True)

        def _run_hub_queue(hub_ip: str, spoke_sites: list[SiteDefinition]) -> None:
            ctx = hub_contexts.get(hub_ip, {})
            if ctx.get("failed") and not args.dry_run:
                with print_lock:
                    print(
                        f"  [Hub {hub_ip}] Hub setup failed — skipping all {len(spoke_sites)} spoke(s).",
                        flush=True,
                    )
                now = clock.now()
                with all_runs_lock:
                    for site in spoke_sites:
                        all_runs.append(SiteRun(site=site, started_at=now, ended_at=now))
                for site in spoke_sites:
                    _emit_progress(f"Skipped '{site.display_name}' (hub setup failed)")
                return

            queue_size = len(spoke_sites)
            for q_index, site in enumerate(spoke_sites, start=1):
                if cancellation_requested():
                    with print_lock:
                        print(
                            f"  [Hub {hub_ip}] Stop requested — halting remaining spoke(s).",
                            flush=True,
                        )
                    break

                # Skip spokes the hub couldn't route to via the server interface.
                if not args.skip_routing_check and not args.dry_run and \
                        ctx.get("routing", {}).get(site.ip_address) is False:
                    with print_lock:
                        print(
                            f"  [Hub {hub_ip}] [{q_index}/{queue_size}] Skipping spoke"
                            f" '{site.ip_address or 'no-ip'}' — no hub route via server"
                            f" interface '{site.hub_server_intf}'.",
                            flush=True,
                        )
                    now = clock.now()
                    rr = next((r for r in ctx.get("routing_results", [])
                               if site.ip_address and site.ip_address in (r.command or "")), None)
                    routing_cmd = CommandResult(
                        template=FORTIGATE_HUB_ROUTING_CHECK,
                        command=(rr.command if rr else f"get router info routing-table details {site.ip_address}"),
                        started_at=now, ended_at=now, return_code=None,
                        stdout=(rr.stdout if rr else ""), stderr="",
                        error=(f"Hub has no route to spoke {site.ip_address} via server "
                               f"interface '{site.hub_server_intf}'."),
                    )
                    site_run = SiteRun(site=site, started_at=now, ended_at=now,
                                       command_results=[routing_cmd])
                    with all_runs_lock:
                        all_runs.append(site_run)
                    _emit_progress(f"Skipped '{site.display_name}' (hub routing failed)")
                    continue

                with print_lock:
                    print(
                        f"  [Hub {hub_ip}] [{q_index}/{queue_size}] Discovering name"
                        f" ({site.ip_address or 'no-ip'})",
                        flush=True,
                    )
                name_result, discovered_name = discover_firewall_name(
                    site, args.firewall_name_command,
                    timeout=args.firewall_name_timeout, dry_run=args.dry_run,
                )
                if discovered_name:
                    set_site_display_name(site, discovered_name)
                elif not args.dry_run:
                    with print_lock:
                        print(
                            f"  [Hub {hub_ip}] Could not discover firewall name;"
                            f" using fallback '{site.display_name}'.",
                            flush=True,
                        )

                if name_result.error and not args.dry_run:
                    with print_lock:
                        print(
                            f"  [Hub {hub_ip}] [{q_index}/{queue_size}] Skipping spoke"
                            f" '{site.display_name}' ({site.ip_address or 'no-ip'})"
                            f" — SSH connection failed: {name_result.error}",
                            flush=True,
                        )
                    now = clock.now()
                    site_run = SiteRun(
                        site=site,
                        started_at=now,
                        ended_at=now,
                        name_discovery_result=name_result,
                    )
                else:
                    with print_lock:
                        print(
                            f"  [Hub {hub_ip}] [{q_index}/{queue_size}] Running spoke"
                            f" '{site.display_name}' ({site.ip_address or 'no-ip'})",
                            flush=True,
                        )
                    site_run = run_fortigate_spoke_only(site, args, name_discovery_result=name_result)

                if q_index < queue_size and args.delay_seconds:
                    site_run.delayed_after_seconds = args.delay_seconds
                    with print_lock:
                        print(
                            f"  [Hub {hub_ip}] Waiting {args.delay_seconds}s before next spoke...",
                            flush=True,
                        )
                    time.sleep(args.delay_seconds)

                with all_runs_lock:
                    all_runs.append(site_run)
                _emit_progress(f"Completed '{site.display_name}' ({site.ip_address or 'no-ip'})")

        print(f"Running spoke tests across {len(hub_queues)} hub queue(s) in parallel...", flush=True)
        queue_threads = [
            threading.Thread(target=_run_hub_queue, args=(hub_ip, spoke_sites), daemon=True)
            for hub_ip, spoke_sites in hub_queues.items()
        ]
        for t in queue_threads:
            t.start()
        for t in queue_threads:
            t.join()

        # Sort collected runs back to the original CSV/XLSX row order.
        runs = sorted(all_runs, key=lambda r: r.site.index)

        if not args.skip_hub_setup:
            # Finalize all hub servers (stop the process and collect output).
            for _, ctx in hub_contexts.items():
                proc = ctx["server_process"]
                if proc is not None and ctx["server_initial"] is not None:
                    ctx["server_result"] = finalize_background_command(
                        ctx["server_initial"], proc, stop_if_running=True
                    )
                    ctx["server_process"] = None
                else:
                    ctx["server_result"] = ctx.get("server_initial")

    else:
        # Custom command mode: discover hub names once per unique hub IP.
        seen_hub_ips: dict[str, str] = {}
        for site in sites:
            if site.hub_ip and site.hub_ip not in seen_hub_ips:
                ssh_target = site.hub_mgmt_ip or site.hub_ip
                hub_site = SiteDefinition(
                    index=0, raw={},
                    placeholders={"spoke_ip": ssh_target},
                    display_name=ssh_target,
                    ip_address=ssh_target,
                    hub_ip=site.hub_ip,
                    speed="", speed_mbps=None,
                    speed_with_margin_mbps=None,
                    speed_with_margin_label="",
                )
                _, hub_name = discover_firewall_name(
                    hub_site, args.firewall_name_command,
                    timeout=args.firewall_name_timeout, dry_run=args.dry_run,
                )
                seen_hub_ips[site.hub_ip] = hub_name or ""
        for site in sites:
            site.hub_name = seen_hub_ips.get(site.hub_ip, "")

        # Run sites sequentially.
        for index, site in enumerate(sites, start=1):
            if cancellation_requested():
                print("Stop requested — halting remaining site(s).", flush=True)
                break
            print(f"[{index}/{total_sites}] Discovering firewall name ({site.ip_address or 'no-ip'})", flush=True)
            name_result, discovered_name = discover_firewall_name(
                site,
                args.firewall_name_command,
                timeout=args.firewall_name_timeout,
                dry_run=args.dry_run,
            )
            if discovered_name:
                set_site_display_name(site, discovered_name)
            elif not args.dry_run:
                print(f"  Could not discover firewall name; using fallback '{site.display_name}'.", flush=True)

            print(f"[{index}/{total_sites}] Running site '{site.display_name}' ({site.ip_address or 'no-ip'})", flush=True)
            site_run = run_site(
                site,
                command_templates,
                timeout=args.timeout,
                dry_run=args.dry_run,
                name_discovery_result=name_result,
            )
            runs.append(site_run)

            if index < total_sites and args.delay_seconds:
                site_run.delayed_after_seconds = args.delay_seconds
                print(f"Waiting {args.delay_seconds} seconds before the next site...", flush=True)
                time.sleep(args.delay_seconds)

    summary = summarize(runs)
    report_html = build_html_report(
        input_path=input_path,
        output_path=output_path,
        results=runs,
        command_templates=active_command_templates,
        delay_seconds=args.delay_seconds,
    )
    output_path.write_text(report_html, encoding="utf-8")
    print(f"HTML report written to: {output_path}")

    build_excel_report(runs, summary, output_path.with_suffix(".xlsx"))
    build_pdf_report(runs, summary, output_path.with_suffix(".pdf"))
    print(
        f"Sites: {summary['total_sites']}, success: {summary['successful_sites']}, "
        f"failed: {summary['failed_sites']}"
    )
    return 0 if summary["failed_sites"] == 0 else 1


def main() -> int:
    parser = build_argument_parser()
    args = parser.parse_args()

    if len(sys.argv) == 1:
        prompt_interactive_inputs(args)
        return _show_progress_window(lambda: _run_tests(args, parser))

    return _run_tests(args, parser)


if __name__ == "__main__":
    sys.exit(main())
