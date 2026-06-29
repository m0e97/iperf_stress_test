"""ISP Compliance Report.

Aggregates historic per-site test results for every device of a chosen ISP over a
time window and scores each test against a configurable SLA percentage of the
device's contracted speed. Produces the data the UI renders plus standalone
HTML / Excel / PDF exports for contract-renewal negotiations.
"""
from __future__ import annotations

import html
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import clock
import main as engine
from webapp import db


# --- computation ----------------------------------------------------------

def compute_isp_report(isp: str, start: datetime, end: datetime, sla_pct: float) -> dict[str, Any]:
    """Build the report structure for one ISP over [start, end) at the given SLA %.

    A test "meets" the SLA when its sender throughput >= sla_pct% of the device's
    contracted speed. Devices with no contracted speed are reported but left
    unscored (compliance shown as N/A).
    """
    start_iso = start.isoformat(timespec="seconds")
    end_iso = end.isoformat(timespec="seconds")
    rows = db.isp_run_rows(isp, start_iso, end_iso)
    devices = db.devices_for_isp(isp)
    ratio = sla_pct / 100.0

    buckets: dict[int, dict[str, Any]] = {}
    for d in devices:
        buckets[d["id"]] = {"device": d, "tests": [], "throughputs": []}
    for r in rows:
        b = buckets.get(r["device_id"])
        if b is None:
            continue
        b["tests"].append(r)
        if r["throughput_mbps"] is not None:
            b["throughputs"].append(r["throughput_mbps"])

    device_reports: list[dict[str, Any]] = []
    threshold_by_dev: dict[int, float | None] = {}
    for b in buckets.values():
        d = b["device"]
        contract = engine.parse_speed_to_mbps(d.get("speed") or "")
        threshold = round(contract * ratio, 2) if contract is not None else None
        threshold_by_dev[d["id"]] = threshold
        met = not_met = 0
        for t in b["tests"]:
            tp = t["throughput_mbps"]
            if tp is None or threshold is None:
                continue
            if tp >= threshold:
                met += 1
            else:
                not_met += 1
        scored = met + not_met
        tps = b["throughputs"]
        avg_tp = round(sum(tps) / len(tps), 2) if tps else None
        device_reports.append({
            "device_id": d["id"],
            "name": d.get("name") or "",
            "spoke_ip": d.get("spoke_ip") or "",
            "circuit_id": d.get("circuit_id") or "",
            "speed": d.get("speed") or "",
            "contract_mbps": contract,
            "threshold_mbps": threshold,
            "total_tests": scored,
            "met": met,
            "not_met": not_met,
            "compliance_pct": round(met / scored * 100, 1) if scored else None,
            "min_mbps": round(min(tps), 2) if tps else None,
            "avg_mbps": avg_tp,
            "max_mbps": round(max(tps), 2) if tps else None,
            "avg_pct_of_contract": round(avg_tp / contract * 100, 1) if (avg_tp is not None and contract) else None,
            "last_test": max((t["started_at"] for t in b["tests"]), default=None),
            "contract_known": contract is not None,
        })

    # Worst compliance first (devices with data), then unscored/no-data devices.
    device_reports.sort(key=lambda d: (d["compliance_pct"] is None, d["compliance_pct"] if d["compliance_pct"] is not None else 0, -d["total_tests"]))

    total_met = sum(d["met"] for d in device_reports)
    total_not_met = sum(d["not_met"] for d in device_reports)
    total_scored = total_met + total_not_met

    return {
        "isp": isp,
        "start": start,
        "end": end,
        "sla_pct": sla_pct,
        "generated_at": clock.now(),
        "devices": device_reports,
        "total_devices": len(device_reports),
        "devices_with_data": sum(1 for d in device_reports if d["total_tests"] > 0),
        "total_tests": total_scored,
        "total_met": total_met,
        "total_not_met": total_not_met,
        "overall_compliance_pct": round(total_met / total_scored * 100, 1) if total_scored else None,
        "worst_performers": [d for d in device_reports if d["compliance_pct"] is not None][:5],
        "trend": _compute_trend(rows, threshold_by_dev, start, end),
    }


def _trend_buckets(start: datetime, end: datetime, target: int = 6) -> list[tuple[datetime, datetime]]:
    """Split [start, end) into up to `target` contiguous sub-periods."""
    span = (end - start).total_seconds()
    if span <= 0:
        return [(start, end)]
    n = min(target, max(1, (end - start).days or 1))
    step = span / n
    out = []
    for i in range(n):
        b0 = start + timedelta(seconds=step * i)
        b1 = end if i == n - 1 else start + timedelta(seconds=step * (i + 1))
        out.append((b0, b1))
    return out


def _compute_trend(rows, threshold_by_dev, start, end) -> list[dict[str, Any]]:
    """Per sub-period compliance across all devices, to show improvement/decline."""
    buckets = _trend_buckets(start, end)
    out = []
    for b0, b1 in buckets:
        b0_iso, b1_iso = b0.isoformat(timespec="seconds"), b1.isoformat(timespec="seconds")
        met = not_met = 0
        for r in rows:
            ts = r["started_at"]
            if ts is None or not (b0_iso <= ts < b1_iso):
                continue
            thr = threshold_by_dev.get(r["device_id"])
            tp = r["throughput_mbps"]
            if tp is None or thr is None:
                continue
            if tp >= thr:
                met += 1
            else:
                not_met += 1
        scored = met + not_met
        out.append({
            "label": b0.strftime("%m-%d") + "–" + b1.strftime("%m-%d"),
            "met": met,
            "not_met": not_met,
            "total": scored,
            "compliance_pct": round(met / scored * 100, 1) if scored else None,
        })
    return out


def compute_all_isps(start: datetime, end: datetime, sla_pct: float) -> dict[str, Any]:
    """Rank every ISP against each other over the same window/SLA for a fleet overview."""
    summaries = []
    for entry in db.list_isps():
        rep = compute_isp_report(entry["isp"], start, end, sla_pct)
        summaries.append({
            "isp": entry["isp"],
            "device_count": entry["device_count"],
            "devices_with_data": rep["devices_with_data"],
            "total_tests": rep["total_tests"],
            "total_met": rep["total_met"],
            "total_not_met": rep["total_not_met"],
            "overall_compliance_pct": rep["overall_compliance_pct"],
        })
    # Best compliance first; ISPs with no data sink to the bottom.
    summaries.sort(key=lambda s: (s["overall_compliance_pct"] is None, -(s["overall_compliance_pct"] or 0)))
    grand_met = sum(s["total_met"] for s in summaries)
    grand_not_met = sum(s["total_not_met"] for s in summaries)
    grand_scored = grand_met + grand_not_met
    return {
        "start": start,
        "end": end,
        "sla_pct": sla_pct,
        "generated_at": clock.now(),
        "isps": summaries,
        "total_isps": len(summaries),
        "total_tests": grand_scored,
        "total_met": grand_met,
        "total_not_met": grand_not_met,
        "overall_compliance_pct": round(grand_met / grand_scored * 100, 1) if grand_scored else None,
    }


# --- helpers ---------------------------------------------------------------

def _fmt(v, suffix: str = "") -> str:
    return "—" if v is None else f"{v}{suffix}"


def _pct_class(pct, sla_pct) -> str:
    if pct is None:
        return "na"
    return "ok" if pct >= sla_pct else ("warn" if pct >= sla_pct * 0.8 else "bad")


def _window_label(report: dict[str, Any]) -> str:
    return f'{report["start"].strftime("%Y-%m-%d")} → {report["end"].strftime("%Y-%m-%d")}'


# --- HTML export (standalone) ---------------------------------------------

def render_html(report: dict[str, Any]) -> str:
    isp = html.escape(report["isp"])
    sla = report["sla_pct"]
    overall = report["overall_compliance_pct"]
    overall_cls = _pct_class(overall, sla)

    rows_html = []
    for d in report["devices"]:
        cls = _pct_class(d["compliance_pct"], sla)
        name = html.escape(d["name"] or d["spoke_ip"])
        rows_html.append(
            f"<tr>"
            f"<td>{name}<div class='sub'>{html.escape(d['spoke_ip'])}</div></td>"
            f"<td>{html.escape(d['circuit_id']) or '—'}</td>"
            f"<td>{html.escape(d['speed']) or '—'}</td>"
            f"<td class='num'>{_fmt(d['threshold_mbps'], ' Mbps')}</td>"
            f"<td class='num'>{d['total_tests']}</td>"
            f"<td class='num ok-t'>{d['met']}</td>"
            f"<td class='num bad-t'>{d['not_met']}</td>"
            f"<td class='num'><span class='pill {cls}'>{_fmt(d['compliance_pct'], '%')}</span></td>"
            f"<td class='num'>{_fmt(d['avg_mbps'])}</td>"
            f"<td class='num'>{_fmt(d['min_mbps'])} / {_fmt(d['max_mbps'])}</td>"
            f"<td class='num'>{_fmt(d['avg_pct_of_contract'], '%')}</td>"
            f"</tr>"
        )

    trend = report.get("trend") or []
    trend_html = ""
    if any(b["total"] for b in trend):
        bars = []
        for b in trend:
            pct = b["compliance_pct"]
            h = int(pct) if pct is not None else 0
            cls = _pct_class(pct, sla)
            lbl = _fmt(pct, "%") if pct is not None else "—"
            bars.append(
                f"<div class='tbar'><div class='tbar-track'><div class='tbar-fill {cls}' style='height:{h}%'></div></div>"
                f"<div class='tval'>{lbl}</div><div class='tlbl'>{html.escape(b['label'])}</div></div>"
            )
        trend_html = ("<h2 class='sec'>Compliance trend</h2>"
                      "<div class='trend'>" + "".join(bars) + "</div>")

    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>ISP Compliance Report — {isp}</title>
<style>
  :root {{ --bg:#f4f6f9; --panel:#fff; --ink:#1a1f2e; --muted:#5a6478; --border:#dde2ec;
           --ok:#0e7a4a; --warn:#b8860b; --bad:#c0392b; --accent:#8c3d2b; --accent-dark:#6e2f20; }}
  body {{ margin:0; background:var(--bg); color:var(--ink); font:14px/1.55 -apple-system,Segoe UI,Roboto,sans-serif; }}
  .wrap {{ max-width:1100px; margin:0 auto; padding:28px; }}
  h1 {{ margin:0 0 4px; font-size:1.5rem; color:var(--accent-dark); }}
  .meta {{ color:var(--muted); margin-bottom:20px; }}
  .cards {{ display:flex; gap:14px; flex-wrap:wrap; margin-bottom:22px; }}
  .card {{ background:var(--panel); border:1px solid var(--border); border-radius:10px; padding:14px 18px; min-width:150px; }}
  .card .label {{ color:var(--muted); font-size:.8rem; text-transform:uppercase; letter-spacing:.05em; }}
  .card .value {{ font-size:1.7rem; font-weight:700; }}
  table {{ width:100%; border-collapse:collapse; background:var(--panel); border:1px solid var(--border); border-radius:10px; overflow:hidden; }}
  th,td {{ padding:9px 12px; border-bottom:1px solid var(--border); text-align:left; }}
  th {{ background:#eef2f7; font-size:.74rem; text-transform:uppercase; letter-spacing:.04em; color:var(--muted); }}
  td.num, th.num {{ text-align:right; }}
  .sub {{ color:var(--muted); font-size:.8rem; }}
  .ok-t {{ color:var(--ok); }} .bad-t {{ color:var(--bad); }}
  .pill {{ display:inline-block; padding:2px 9px; border-radius:20px; font-weight:700; font-size:.85rem; }}
  .pill.ok {{ background:#dcf3e8; color:var(--ok); }}
  .pill.warn {{ background:#fbf0d5; color:var(--warn); }}
  .pill.bad {{ background:#f8dcd8; color:var(--bad); }}
  .pill.na {{ background:#eceff3; color:var(--muted); }}
  .value.ok {{ color:var(--ok); }} .value.warn {{ color:var(--warn); }} .value.bad {{ color:var(--bad); }}
  .note {{ color:var(--muted); font-size:.85rem; margin-top:16px; }}
  h2.sec {{ font-size:1rem; margin:18px 0 10px; color:var(--accent-dark); }}
  .trend {{ display:flex; gap:10px; align-items:flex-end; background:var(--panel); border:1px solid var(--border); border-radius:10px; padding:16px; margin-bottom:22px; }}
  .tbar {{ flex:1; display:flex; flex-direction:column; align-items:center; gap:4px; }}
  .tbar-track {{ width:100%; max-width:60px; height:90px; background:#eef2f7; border-radius:5px; display:flex; align-items:flex-end; overflow:hidden; }}
  .tbar-fill {{ width:100%; border-radius:5px 5px 0 0; }}
  .tbar-fill.ok {{ background:var(--ok); }} .tbar-fill.warn {{ background:var(--warn); }} .tbar-fill.bad {{ background:var(--bad); }} .tbar-fill.na {{ background:#cbd2dc; }}
  .tval {{ font-weight:700; font-size:.82rem; }} .tlbl {{ color:var(--muted); font-size:.72rem; }}
</style></head><body><div class="wrap">
  <h1>ISP Compliance Report — {isp}</h1>
  <div class="meta">Window {_window_label(report)} · SLA target {sla:g}% of contracted speed · generated {report['generated_at'].strftime('%Y-%m-%d %H:%M')}</div>
  <div class="cards">
    <div class="card"><div class="label">Overall compliance</div><div class="value {overall_cls}">{_fmt(overall, '%')}</div></div>
    <div class="card"><div class="label">Devices</div><div class="value">{report['devices_with_data']}<span style="font-size:1rem;color:var(--muted)"> / {report['total_devices']}</span></div></div>
    <div class="card"><div class="label">Tests scored</div><div class="value">{report['total_tests']}</div></div>
    <div class="card"><div class="label">Met</div><div class="value ok">{report['total_met']}</div></div>
    <div class="card"><div class="label">Not met</div><div class="value bad">{report['total_not_met']}</div></div>
  </div>
  {trend_html}
  <table>
    <thead><tr>
      <th>Device</th><th>Circuit</th><th>Contract</th><th class="num">SLA threshold</th>
      <th class="num">Tests</th><th class="num">Met</th><th class="num">Not met</th>
      <th class="num">Compliance</th><th class="num">Avg Mbps</th><th class="num">Min / Max</th><th class="num">Avg % of contract</th>
    </tr></thead>
    <tbody>{''.join(rows_html)}</tbody>
  </table>
  <p class="note">Devices are ordered worst-compliance first. "Met" = sender throughput ≥ {sla:g}% of the device's contracted speed.
  Devices with no contracted speed configured are shown with N/A compliance. Devices listed with 0 tests had no runs in this window.</p>
</div></body></html>"""


# --- Excel export ----------------------------------------------------------

def build_excel(report: dict[str, Any], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "ISP Compliance"
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F2532")
    head_font = Font(bold=True, color="FFFFFF")

    ws.append([f"ISP Compliance Report — {report['isp']}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Window {_window_label(report)}  |  SLA target {report['sla_pct']:g}% of contracted speed"])
    ws.append([f"Generated {report['generated_at'].strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Overall compliance", _fmt(report["overall_compliance_pct"], "%"),
               "Devices (with data / total)", f"{report['devices_with_data']} / {report['total_devices']}",
               "Tests", report["total_tests"], "Met", report["total_met"], "Not met", report["total_not_met"]])
    ws.append([])

    headers = ["Device", "Spoke IP", "Circuit", "Contract", "SLA threshold (Mbps)",
               "Tests", "Met", "Not met", "Compliance %", "Avg Mbps", "Min Mbps", "Max Mbps",
               "Avg % of contract", "Last test"]
    hrow = ws.max_row + 1
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.fill = head_fill
        cell.font = head_font

    for d in report["devices"]:
        ws.append([
            d["name"] or d["spoke_ip"], d["spoke_ip"], d["circuit_id"], d["speed"],
            d["threshold_mbps"], d["total_tests"], d["met"], d["not_met"],
            d["compliance_pct"], d["avg_mbps"], d["min_mbps"], d["max_mbps"],
            d["avg_pct_of_contract"], (d["last_test"] or "")[:19].replace("T", " "),
        ])
    widths = [22, 15, 14, 12, 18, 8, 8, 9, 13, 11, 11, 11, 17, 19]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    trend = [b for b in (report.get("trend") or []) if b["total"]]
    if trend:
        ws.append([])
        trow = ws.max_row + 1
        ws.append(["Compliance trend"] + [b["label"] for b in trend])
        ws.cell(row=trow, column=1).font = bold
        ws.append(["Compliance %"] + [b["compliance_pct"] for b in trend])
        ws.append(["Met / scored"] + [f"{b['met']}/{b['total']}" for b in trend])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# --- PDF export ------------------------------------------------------------

def build_pdf(report: dict[str, Any], output_path: Path) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], textColor=colors.HexColor("#6e2f20"), fontSize=18)
    sub = ParagraphStyle("s", parent=styles["Normal"], textColor=colors.HexColor("#5a6478"), fontSize=9)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(output_path), pagesize=landscape(A4),
                            leftMargin=1.2 * cm, rightMargin=1.2 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    story = [
        Paragraph(f"ISP Compliance Report — {html.escape(report['isp'])}", title),
        Paragraph(f"Window {_window_label(report)} &nbsp;|&nbsp; SLA target {report['sla_pct']:g}% of contracted speed "
                  f"&nbsp;|&nbsp; generated {report['generated_at'].strftime('%Y-%m-%d %H:%M')}", sub),
        Spacer(1, 8),
        Paragraph(f"<b>Overall compliance: {_fmt(report['overall_compliance_pct'], '%')}</b> &nbsp; "
                  f"Devices with data: {report['devices_with_data']}/{report['total_devices']} &nbsp; "
                  f"Tests: {report['total_tests']} &nbsp; Met: {report['total_met']} &nbsp; Not met: {report['total_not_met']}",
                  styles["Normal"]),
        Spacer(1, 10),
    ]

    data = [["Device", "Circuit", "Contract", "SLA thr.", "Tests", "Met", "Not met", "Compl.%", "Avg", "Min", "Max", "% contr."]]
    for d in report["devices"]:
        data.append([
            (d["name"] or d["spoke_ip"]), d["circuit_id"] or "—", d["speed"] or "—",
            _fmt(d["threshold_mbps"]), str(d["total_tests"]), str(d["met"]), str(d["not_met"]),
            _fmt(d["compliance_pct"]), _fmt(d["avg_mbps"]), _fmt(d["min_mbps"]), _fmt(d["max_mbps"]),
            _fmt(d["avg_pct_of_contract"]),
        ])
    tbl = Table(data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E2532")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dde2ec")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9fb")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    # Colour compliance cell by SLA outcome.
    sla = report["sla_pct"]
    for i, d in enumerate(report["devices"], start=1):
        c = d["compliance_pct"]
        if c is None:
            continue
        col = colors.HexColor("#0e7a4a") if c >= sla else (colors.HexColor("#b8860b") if c >= sla * 0.8 else colors.HexColor("#c0392b"))
        style.append(("TEXTCOLOR", (7, i), (7, i), col))
        style.append(("FONTNAME", (7, i), (7, i), "Helvetica-Bold"))
    tbl.setStyle(TableStyle(style))
    story.append(tbl)

    trend = [b for b in (report.get("trend") or []) if b["total"]]
    if trend:
        story.append(Spacer(1, 14))
        story.append(Paragraph("<b>Compliance trend</b>", styles["Normal"]))
        story.append(Spacer(1, 4))
        tdata = [["Sub-period"] + [b["label"] for b in trend],
                 ["Compliance %"] + [_fmt(b["compliance_pct"]) for b in trend],
                 ["Met / scored"] + [f"{b['met']}/{b['total']}" for b in trend]]
        ttbl = Table(tdata, repeatRows=1)
        ttbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E2532")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dde2ec")),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ]))
        story.append(ttbl)
    doc.build(story)


# --- All-ISPs comparison renderers ----------------------------------------

def render_all_html(summary: dict[str, Any]) -> str:
    sla = summary["sla_pct"]
    rows = []
    for s in summary["isps"]:
        cls = _pct_class(s["overall_compliance_pct"], sla)
        rows.append(
            f"<tr><td>{html.escape(s['isp'])}</td>"
            f"<td class='num'>{s['devices_with_data']} / {s['device_count']}</td>"
            f"<td class='num'>{s['total_tests']}</td>"
            f"<td class='num ok-t'>{s['total_met']}</td>"
            f"<td class='num bad-t'>{s['total_not_met']}</td>"
            f"<td class='num'><span class='pill {cls}'>{_fmt(s['overall_compliance_pct'], '%')}</span></td></tr>"
        )
    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8"><title>ISP Compliance — All ISPs</title>
<style>
  :root {{ --bg:#f4f6f9; --panel:#fff; --ink:#1a1f2e; --muted:#5a6478; --border:#dde2ec;
           --ok:#0e7a4a; --warn:#b8860b; --bad:#c0392b; --accent-dark:#6e2f20; }}
  body {{ margin:0; background:var(--bg); color:var(--ink); font:14px/1.55 -apple-system,Segoe UI,Roboto,sans-serif; }}
  .wrap {{ max-width:820px; margin:0 auto; padding:28px; }}
  h1 {{ margin:0 0 4px; font-size:1.5rem; color:var(--accent-dark); }}
  .meta {{ color:var(--muted); margin-bottom:20px; }}
  table {{ width:100%; border-collapse:collapse; background:var(--panel); border:1px solid var(--border); border-radius:10px; overflow:hidden; }}
  th,td {{ padding:9px 12px; border-bottom:1px solid var(--border); text-align:left; }}
  th {{ background:#eef2f7; font-size:.74rem; text-transform:uppercase; letter-spacing:.04em; color:var(--muted); }}
  td.num, th.num {{ text-align:right; }}
  .ok-t {{ color:var(--ok); }} .bad-t {{ color:var(--bad); }}
  .pill {{ display:inline-block; padding:2px 9px; border-radius:20px; font-weight:700; font-size:.85rem; }}
  .pill.ok {{ background:#dcf3e8; color:var(--ok); }} .pill.warn {{ background:#fbf0d5; color:var(--warn); }}
  .pill.bad {{ background:#f8dcd8; color:var(--bad); }} .pill.na {{ background:#eceff3; color:var(--muted); }}
</style></head><body><div class="wrap">
  <h1>ISP Compliance — All ISPs</h1>
  <div class="meta">Window {_window_label(summary)} · SLA target {sla:g}% · overall {_fmt(summary['overall_compliance_pct'], '%')} · generated {summary['generated_at'].strftime('%Y-%m-%d %H:%M')}</div>
  <table><thead><tr><th>ISP</th><th class="num">Devices</th><th class="num">Tests</th><th class="num">Met</th><th class="num">Not met</th><th class="num">Compliance</th></tr></thead>
  <tbody>{''.join(rows)}</tbody></table>
</div></body></html>"""


def build_all_excel(summary: dict[str, Any], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "ISP Comparison"
    ws.append(["ISP Compliance — All ISPs"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Window {_window_label(summary)}  |  SLA target {summary['sla_pct']:g}%  |  Overall {_fmt(summary['overall_compliance_pct'], '%')}"])
    ws.append([])
    headers = ["ISP", "Devices with data", "Device count", "Tests", "Met", "Not met", "Compliance %"]
    hrow = ws.max_row + 1
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.fill = PatternFill("solid", fgColor="1F2532")
        cell.font = Font(bold=True, color="FFFFFF")
    for s in summary["isps"]:
        ws.append([s["isp"], s["devices_with_data"], s["device_count"], s["total_tests"],
                   s["total_met"], s["total_not_met"], s["overall_compliance_pct"]])
    for i, w in enumerate([22, 18, 14, 10, 8, 9, 14], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_all_pdf(summary: dict[str, Any], output_path: Path) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], textColor=colors.HexColor("#6e2f20"), fontSize=18)
    sub = ParagraphStyle("s", parent=styles["Normal"], textColor=colors.HexColor("#5a6478"), fontSize=9)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(output_path), pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    sla = summary["sla_pct"]
    story = [
        Paragraph("ISP Compliance — All ISPs", title),
        Paragraph(f"Window {_window_label(summary)} &nbsp;|&nbsp; SLA target {sla:g}% &nbsp;|&nbsp; "
                  f"overall {_fmt(summary['overall_compliance_pct'], '%')} &nbsp;|&nbsp; generated {summary['generated_at'].strftime('%Y-%m-%d %H:%M')}", sub),
        Spacer(1, 12),
    ]
    data = [["ISP", "Devices", "Tests", "Met", "Not met", "Compliance %"]]
    for s in summary["isps"]:
        data.append([s["isp"], f"{s['devices_with_data']}/{s['device_count']}", str(s["total_tests"]),
                     str(s["total_met"]), str(s["total_not_met"]), _fmt(s["overall_compliance_pct"])])
    tbl = Table(data, repeatRows=1)
    st = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E2532")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dde2ec")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9fb")]),
    ]
    for i, s in enumerate(summary["isps"], start=1):
        c = s["overall_compliance_pct"]
        if c is None:
            continue
        col = colors.HexColor("#0e7a4a") if c >= sla else (colors.HexColor("#b8860b") if c >= sla * 0.8 else colors.HexColor("#c0392b"))
        st.append(("TEXTCOLOR", (5, i), (5, i), col))
        st.append(("FONTNAME", (5, i), (5, i), "Helvetica-Bold"))
    tbl.setStyle(TableStyle(st))
    story.append(tbl)
    doc.build(story)
